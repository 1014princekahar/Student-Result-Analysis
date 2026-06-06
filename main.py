import sys
import os

# Fix Windows CP1252 emoji encoding issue globally
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

import sqlite3
import pandas as pd
import time
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QPushButton, QLabel, QStackedWidget, 
                               QListWidget, QListWidgetItem, QFrame, QFileDialog, 
                               QTableWidget, QTableWidgetItem, QProgressBar, 
                               QMessageBox, QComboBox, QHeaderView, QGraphicsDropShadowEffect, 
                               QGroupBox, QSizePolicy, QLineEdit, QSpinBox, QTabWidget, QLabel,
                               QScrollArea, QAbstractItemView, QDialog, QButtonGroup, QRadioButton
                              )
from PySide6.QtCore import Qt, QThread, Signal, QTimer, QStandardPaths, QSize, QMargins, QPropertyAnimation, QEasingCurve
from PySide6.QtGui import QFont, QIcon, QColor, QPixmap, QCursor, QKeySequence, QShortcut, QPainter, QPen, QBrush, QFontMetrics
from PySide6.QtCore import QDate, QRect
from PySide6.QtWidgets import QSplashScreen
from PySide6.QtCore import QObject
try:
    from PySide6.QtPrintSupport import QPrinter, QPrintDialog
    PRINT_SUPPORT = True
except ImportError:
    PRINT_SUPPORT = False

import matplotlib
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table as PdfTable, TableStyle, Paragraph, Spacer

# ==============================================================================
# 1. PATHS & ASSETS
# ==============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(BASE_DIR, 'assets')
LOGO_PATH = os.path.join(ASSETS_DIR, 'logo.png')
ICON_PATH = os.path.join(ASSETS_DIR, 'icon.ico')
SASCMA_PDF_HEADER_PATH = os.path.join(ASSETS_DIR, 'sascma_pdf_header.png')
UNIVERSITY_PDF_HEADER_PATH = os.path.join(ASSETS_DIR, 'university_pdf_header.png')
PDF_HEADER_PATH = UNIVERSITY_PDF_HEADER_PATH
COURSE_NAME = "SASCMA – STERS  |  VNSGU"

# ==============================================================================
# 2. IMPORT PARSING LOGIC (Robust Import)
# ==============================================================================
PARSER_LOADED = False
try:
    import core.pdf_parser as final_result
    print("Loaded core.pdf parser module.")
    PARSER_LOADED = True
except ImportError:
    try:
        import core.pdf_parser as final_result
        print("Loaded core.pdf_parser module.")
        PARSER_LOADED = True
    except ImportError:
        final_result = None
        print("CRITICAL: Parsing logic missing. Ensure final_result.py is in the root folder.")
    # Don't exit, app will show a message in the UI

matplotlib.use('QtAgg')

# ==============================================================================
# 3. STYLESHEET  — VNSGU Portal Inspired
# ==============================================================================
VNSGU_STYLESHEET = """
/* ===== GLOBAL ===== */
QWidget {
    background-color: #f0f4f8;
    color: #1e293b;
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    font-size: 13px;
}

/* ===== SIDEBAR — Deep Navy ===== */
QFrame#sidebar {
    background-color: #0c1a3a;
    border-right: 3px solid #f97316;
    border-radius: 0px;
}
QFrame#sidebar_logo_container {
    background-color: #1a2e5a;
    border-bottom: 1px solid rgba(255,255,255,0.08);
}
QLabel#sidebar_college_name {
    color: #ffffff;
    font-size: 14px;
    font-weight: 700;
    background: transparent;
    padding: 2px 8px;
    letter-spacing: 0.3px;
}
QPushButton#toggle_btn {
    background-color: rgba(249, 115, 22, 0.15);
    color: #f97316;
    border: 1px solid rgba(249, 115, 22, 0.4);
    border-radius: 4px;
    padding: 4px 10px;
    font-weight: 700;
    font-size: 11px;
}
QPushButton#toggle_btn:hover {
    background-color: rgba(249, 115, 22, 0.3);
}
QLabel#sidebar_logo {
    background-color: transparent;
    border-radius: 50%;
    border: 2px solid transparent;
}
QLabel#sidebar_logo:hover {
    border: 2px solid rgba(249,115,22,0.5);
}

/* ===== FULL MENU LIST ===== */
QListWidget#menu_list {
    background-color: transparent;
    border: none;
    outline: none;
    padding: 6px 0px;
}
QListWidget#menu_list::item {
    color: #94a3b8;
    padding: 13px 20px;
    border-left: 4px solid transparent;
    font-size: 13px;
    font-weight: 500;
    margin: 1px 0px;
}
QListWidget#menu_list::item:hover {
    background-color: rgba(255,255,255,0.06);
    color: #e2e8f0;
    border-left: 4px solid rgba(249,115,22,0.4);
}
QListWidget#menu_list::item:selected {
    background-color: #1a2e5a;
    color: #ffffff;
    border-left: 4px solid #f97316;
    font-weight: 700;
}

/* ===== ICON-ONLY COLLAPSED LIST ===== */
QListWidget#icon_menu_list {
    background-color: transparent;
    border: none;
    outline: none;
    padding: 6px 0px;
}
QListWidget#icon_menu_list::item {
    color: #94a3b8;
    padding: 10px 0px;
    font-size: 20px;
    text-align: center;
    margin: 2px 4px;
    border-radius: 6px;
    border-left: 3px solid transparent;
}
QListWidget#icon_menu_list::item:hover {
    background-color: rgba(255,255,255,0.08);
    color: #f97316;
    border-left: 3px solid rgba(249,115,22,0.4);
}
QListWidget#icon_menu_list::item:selected {
    background-color: #1a2e5a;
    color: #f97316;
    border-left: 3px solid #f97316;
}

/* ===== SETTINGS BUTTON ===== */
QPushButton#settings_btn {
    background-color: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.12);
    color: #94a3b8;
    margin: 8px 14px 14px 14px;
    text-align: left;
    padding: 11px 18px;
    border-radius: 6px;
    font-size: 13px;
}
QPushButton#settings_btn:hover {
    background-color: rgba(255,255,255,0.12);
    color: #ffffff;
}
QPushButton#icon_settings_btn {
    background-color: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.12);
    color: #94a3b8;
    margin: 8px 6px 14px 6px;
    padding: 10px 0px;
    border-radius: 6px;
    font-size: 18px;
}
QPushButton#icon_settings_btn:hover {
    background-color: rgba(255,255,255,0.14);
    color: #f97316;
}

/* ===== TOP HEADER — Teal Banner ===== */
QFrame#main_header {
    background-color: #0d9488;
    border-bottom: 2px solid #0f766e;
    min-height: 55px;
    max-height: 55px;
}
QLabel#main_title {
    color: #ffffff;
    font-size: 18px;
    font-weight: 700;
    letter-spacing: 0.3px;
}
QLabel#header_date {
    color: rgba(255,255,255,0.75);
    font-size: 12px;
    font-weight: 500;
}
QLabel#header_welcome {
    color: #ffffff;
    font-size: 13px;
    font-weight: 600;
    background-color: rgba(255,255,255,0.15);
    border-radius: 4px;
    padding: 4px 12px;
}

/* ===== CARDS ===== */
QFrame#card {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    padding: 18px;
}
QFrame#card_title_bar {
    background-color: #1a2e5a;
    border-radius: 6px 6px 0px 0px;
    padding: 8px 16px;
    min-height: 36px;
    max-height: 36px;
}
QLabel#card_title_label {
    color: #ffffff;
    font-weight: 700;
    font-size: 13px;
}
QFrame#card_collapsible {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
}

/* ===== KPI CARDS ===== */
QFrame#kpi_card {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    border-top: 4px solid #0d9488;
}
QLabel#kpi_value {
    font-size: 30px;
    font-weight: 800;
    color: #0c1a3a;
}
QLabel#kpi_label {
    font-size: 11px;
    color: #64748b;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

/* ===== BUTTONS ===== */
QPushButton {
    background-color: #0d9488;
    color: white;
    border-radius: 5px;
    padding: 9px 18px;
    font-weight: 700;
    border: 1px solid #0f766e;
    font-size: 13px;
    min-height: 20px;
}
QPushButton:hover {
    background-color: #0f766e;
    border-color: #115e59;
}
QPushButton:pressed {
    background-color: #115e59;
    padding-top: 10px;
    padding-bottom: 8px;
}
QPushButton:disabled {
    background-color: #e2e8f0;
    color: #94a3b8;
    border: 1px solid #cbd5e1;
}
QPushButton#secondary {
    background-color: #ffffff;
    border: 1.5px solid #0d9488;
    color: #0d9488;
}
QPushButton#secondary:hover {
    background-color: #f0fdfa;
    border-color: #0f766e;
    color: #0f766e;
}
QPushButton#danger {
    background-color: #dc2626;
    border-color: #b91c1c;
    color: #ffffff;
}
QPushButton#danger:hover { background-color: #b91c1c; }
QPushButton#orange_btn {
    background-color: #f97316;
    border-color: #ea580c;
}
QPushButton#orange_btn:hover { background-color: #ea580c; }

/* ===== TABLE ===== */
QTableWidget {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    gridline-color: #f1f5f9;
    selection-background-color: #ccfbf1;
    selection-color: #0c1a3a;
    alternate-background-color: #f8fafc;
}
QTableWidget::item {
    padding: 4px 8px;
    border-bottom: 1px solid #f1f5f9;
    min-height: 30px;
    color: #1e293b;
}
QTableWidget::item:selected { color: #0c1a3a; background-color: #ccfbf1; }
QHeaderView::section {
    background-color: #f97316;
    color: #ffffff;
    padding: 9px 10px;
    border: none;
    border-right: 1px solid rgba(255,255,255,0.2);
    font-weight: 700;
    font-size: 12px;
    letter-spacing: 0.2px;
}
QHeaderView::section:first {
    border-top-left-radius: 0px;
}
QHeaderView::section:last {
    border-top-right-radius: 0px;
    border-right: none;
}

/* ===== DROP ZONE ===== */
QFrame#drop_zone {
    background-color: #ffffff;
    border: 2px dashed #cbd5e1;
    border-radius: 12px;
}
QFrame#drop_zone:hover {
    border-color: #0d9488;
    background-color: #f0fdfa;
}

/* ===== INPUTS ===== */
QLineEdit {
    padding: 9px 12px;
    border: 1.5px solid #cbd5e1;
    border-radius: 5px;
    background-color: #ffffff;
    color: #1e293b;
    font-size: 13px;
}
QLineEdit:focus {
    border: 1.5px solid #0d9488;
    background-color: #f0fdfa;
}
QComboBox {
    padding: 8px 12px;
    border: 1.5px solid #cbd5e1;
    border-radius: 5px;
    background-color: #ffffff;
    color: #1e293b;
    font-size: 13px;
}
QComboBox:focus { border: 1.5px solid #0d9488; }
QComboBox::drop-down { border: none; width: 28px; }
QSpinBox {
    padding: 8px 10px;
    border: 1.5px solid #cbd5e1;
    border-radius: 5px;
    background-color: #ffffff;
    color: #1e293b;
}
QSpinBox:focus { border: 1.5px solid #0d9488; }

/* ===== SCROLLBAR ===== */
QScrollBar:vertical {
    background: #f1f5f9;
    width: 8px;
    border-radius: 4px;
}
QScrollBar::handle:vertical {
    background: #94a3b8;
    border-radius: 4px;
    min-height: 30px;
}
QScrollBar::handle:vertical:hover { background: #0d9488; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
QScrollBar:horizontal {
    background: #f1f5f9;
    height: 8px;
    border-radius: 4px;
}
QScrollBar::handle:horizontal {
    background: #94a3b8;
    border-radius: 4px;
}
QScrollBar::handle:horizontal:hover { background: #0d9488; }

/* ===== TABS ===== */
QTabWidget::pane {
    border: 1px solid #e2e8f0;
    background-color: #ffffff;
    border-radius: 0px 8px 8px 8px;
}
QTabBar::tab {
    background: #e2e8f0;
    color: #334155;
    font-weight: 600;
    padding: 10px 20px;
    border: 1px solid #e2e8f0;
    border-bottom: none;
    border-radius: 6px 6px 0 0;
    margin-right: 2px;
    font-size: 13px;
}
QTabBar::tab:selected {
    background-color: #0c1a3a;
    color: #ffffff;
}
QTabBar::tab:hover:!selected {
    background-color: #cbd5e1;
    color: #1e293b;
}

/* ===== PROGRESS BAR ===== */
QProgressBar {
    background-color: #e2e8f0;
    border: none;
    border-radius: 5px;
    height: 10px;
    text-align: center;
    color: transparent;
}
QProgressBar::chunk {
    background-color: #0d9488;
    border-radius: 5px;
}

/* ===== FOOTER ===== */
QFrame#footer {
    background-color: #1a2e5a;
    color: #94a3b8;
    padding: 8px 20px;
    max-height: 36px;
    min-height: 36px;
}
QLabel#footer_lbl {
    color: #94a3b8;
    font-size: 11px;
}

/* ===== SECTION HEADERS ===== */
QFrame#section_header {
    background-color: #1a2e5a;
    border-radius: 6px 6px 0px 0px;
    padding: 0px;
    min-height: 38px;
    max-height: 38px;
}
QLabel#section_header_label {
    color: #ffffff;
    font-weight: 700;
    font-size: 13px;
    padding-left: 14px;
}
QPushButton#section_collapse_btn {
    background-color: rgba(255,255,255,0.12);
    border: 1px solid rgba(255,255,255,0.2);
    color: white;
    font-size: 14px;
    font-weight: bold;
    padding: 2px 10px;
    min-height: 22px;
    max-height: 22px;
    border-radius: 3px;
    margin-right: 10px;
}
QPushButton#section_collapse_btn:hover {
    background-color: rgba(255,255,255,0.25);
}

/* ===== NOTICE BOX ===== */
QFrame#notice_box {
    background-color: #fffbeb;
    border: 1px solid #fbbf24;
    border-radius: 6px;
    padding: 10px 16px;
}
QLabel#notice_text {
    color: #92400e;
    font-size: 12px;
    font-style: italic;
}

/* ===== GENERAL LABELS on WHITE bg ===== */
QLabel {
    color: #1e293b;
    background: transparent;
}
QGroupBox {
    color: #1e293b;
    font-weight: 700;
}

/* ===== STATUS BADGE ===== */
QLabel#badge_pass {
    background-color: #16a34a;
    color: white;
    font-weight: 700;
    font-size: 11px;
    padding: 2px 8px;
    border-radius: 3px;
}
QLabel#badge_fail {
    background-color: #dc2626;
    color: white;
    font-weight: 700;
    font-size: 11px;
    padding: 2px 8px;
    border-radius: 3px;
}
QLabel#badge_present {
    background-color: #16a34a;
    color: white;
    font-weight: 700;
    font-size: 11px;
    padding: 2px 8px;
    border-radius: 3px;
}
QLabel#badge_internal {
    background-color: #f97316;
    color: white;
    font-weight: 700;
    font-size: 10px;
    padding: 2px 6px;
    border-radius: 3px;
}
QLabel#badge_level {
    background-color: #0c1a3a;
    color: white;
    font-weight: 700;
    font-size: 10px;
    padding: 2px 7px;
    border-radius: 3px;
}
"""

# ==============================================================================
# 4. DATABASE — Imported from core/database.py
# ==============================================================================
from core.database import DatabaseManager


# ==============================================================================
# 5. PDF WORKER
# ==============================================================================
class PDFWorker(QThread):
    finished = Signal(object)   # emits dict with students + metadata
    error    = Signal(str)
    progress = Signal(int)
    log      = Signal(str)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            self.log.emit("Initializing Parser...")

            if not PARSER_LOADED or final_result is None:
                raise ImportError("Parser module not found.")

            # --- Step 1: Detect subjects & scheme ---
            count = final_result.configure_subject_metadata(self.file_path)
            self.log.emit(f"Detected {count} subjects.")
            self.progress.emit(20)

            # --- Step 2: Extract course/semester/year from PDF header ---
            course_info   = getattr(final_result, 'COURSE_INFO', {})
            academic_year = getattr(final_result, 'ACADEMIC_YEAR', '')
            college_name  = getattr(final_result, 'COLLEGE_NAME', '')
            subjects_map  = getattr(final_result, 'SUBJECTS', {})
            ext_passing   = getattr(final_result, 'EXT_PASSING_MIN', [18,18,18,18,9,9,9])

            # Fallback: extract from PDF if not already populated
            if not course_info:
                try:
                    course_info = final_result.extract_course_info_from_pdf(self.file_path)
                except Exception:
                    course_info = {}

            self.progress.emit(35)

            # --- Step 3: Load students ---
            students = final_result.load_students(self.file_path)
            self.progress.emit(80)

            if not students:
                self.error.emit("No students found. Check PDF format.")
                return

            # Progress update during packaging
            for idx in range(len(students)):
                if idx % 10 == 0:
                    prog = 80 + int((idx / len(students)) * 20)
                    self.progress.emit(min(prog, 99))

            self.progress.emit(100)

            # --- Emit: raw students + all metadata ---
            self.finished.emit({
                "students":      students,
                "subjects_map":  subjects_map,
                "ext_passing":   ext_passing,
                "course":        course_info.get('course', 'Unknown Course'),
                "semester":      course_info.get('semester', 'Unknown Semester'),
                "academic_year": academic_year,
                "college_name":  college_name,
                "pdf_filename":  os.path.basename(self.file_path),
                "max_marks":     700,
            })

        except ImportError:
            self.error.emit("Error: Could not load parsing logic. Ensure 'core/pdf_parser.py' exists.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.error.emit(f"Parsing Failed: {str(e)}")

# ==============================================================================
# 6. CUSTOM WIDGETS
# ==============================================================================
class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=5, height=4.0, dpi=100): 
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.fig = fig
        self.axes = fig.add_subplot(111)
        fig.patch.set_facecolor('#ffffff') 
        super(MplCanvas, self).__init__(fig)

class KPICard(QFrame):
    def __init__(self, title, value, color="#0d9488", icon="📊"):
        super().__init__()
        self.setObjectName("kpi_card")
        layout = QVBoxLayout()
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(4)
        self.setLayout(layout)
        
        top_row = QHBoxLayout()
        icon_lbl = QLabel(icon)
        icon_lbl.setStyleSheet(f"font-size: 20px; color: {color}; background: transparent;")
        top_row.addWidget(icon_lbl)
        top_row.addStretch()
        layout.addLayout(top_row)
        
        self.lbl_value = QLabel(str(value))
        self.lbl_value.setObjectName("kpi_value")
        self.lbl_value.setStyleSheet(f"font-size: 30px; font-weight: 800; color: {color}; background: transparent;")
        
        self.lbl_title = QLabel(title)
        self.lbl_title.setObjectName("kpi_label")
        self.lbl_title.setStyleSheet("background: transparent;")
        
        layout.addWidget(self.lbl_value)
        layout.addWidget(self.lbl_title)


def make_section_card(title, collapse_btn_text="−", collapsible=True):
    """Returns (outer_frame, content_frame, collapse_btn|None)"""
    outer = QFrame()
    outer.setObjectName("card_collapsible")
    outer_layout = QVBoxLayout(outer)
    outer_layout.setContentsMargins(0, 0, 0, 0)
    outer_layout.setSpacing(0)
    
    # Header bar
    header = QFrame()
    header.setObjectName("section_header")
    h_layout = QHBoxLayout(header)
    h_layout.setContentsMargins(0, 0, 0, 0)
    
    lbl = QLabel(title)
    lbl.setObjectName("section_header_label")
    h_layout.addWidget(lbl)
    h_layout.addStretch()
    
    collapse_btn = None
    if collapsible:
        collapse_btn = QPushButton(collapse_btn_text)
        collapse_btn.setObjectName("section_collapse_btn")
        collapse_btn.setFixedSize(26, 26)
        h_layout.addWidget(collapse_btn)
    
    outer_layout.addWidget(header)
    
    # Content — explicit colors so nothing is white-on-white
    content = QFrame()
    content.setObjectName("section_content_card")
    content.setStyleSheet(
        "QFrame#section_content_card { background: #ffffff; border-radius: 0px 0px 6px 6px; }"
        "QLabel { color: #1e293b; background: transparent; }"
        "QSpinBox { color: #1e293b; background: #ffffff; }"
        "QComboBox { color: #1e293b; background: #ffffff; }"
        "QLineEdit { color: #1e293b; background: #ffffff; }"
    )
    content_layout = QVBoxLayout(content)
    content_layout.setContentsMargins(14, 14, 14, 14)
    content_layout.setSpacing(8)
    outer_layout.addWidget(content)
    
    return outer, content, collapse_btn, content_layout


# ==============================================================================
# 7. REPORT TYPE DIALOG
# ==============================================================================
class ReportTypeDialog(QDialog):
    """Settings page par export karte waqt report type aur format select karne ka dialog."""
    def __init__(self, parent, stats, top_n=10):
        super().__init__(parent)
        self.setWindowTitle("Generate Report — Select Options")
        self.setFixedSize(500, 460)
        self.setModal(True)
        self.top_n = top_n
        self.setStyleSheet("""
            QDialog { background: #f0f4f8; }
            QLabel#dlg_title { font-size: 17px; font-weight: 700; color: #0c1a3a; }
            QLabel#dlg_section { font-size: 10px; font-weight: 700; color: #64748b;
                                  letter-spacing: 1.2px; padding: 4px 0px 2px 0px; }
            QRadioButton { font-size: 13px; color: #1e293b; background: transparent;
                           padding: 8px 12px; border-radius: 6px; }
            QRadioButton:hover { background: rgba(13,148,136,0.08); }
            QRadioButton::indicator { width: 18px; height: 18px; }
            QRadioButton::indicator:checked { background: #0d9488; border: 2px solid #0d9488;
                                              border-radius: 9px; }
            QRadioButton::indicator:unchecked { background: #ffffff; border: 2px solid #cbd5e1;
                                                border-radius: 9px; }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(10)

        # Header
        title_row = QHBoxLayout()
        icon_lbl = QLabel("📊")
        icon_lbl.setStyleSheet("font-size: 24px; background: transparent;")
        title_lbl = QLabel("Generate Report")
        title_lbl.setObjectName("dlg_title")
        title_row.addWidget(icon_lbl)
        title_row.addSpacing(8)
        title_row.addWidget(title_lbl)
        title_row.addStretch()
        layout.addLayout(title_row)

        div = QFrame(); div.setFixedHeight(1)
        div.setStyleSheet("background: #e2e8f0; margin: 4px 0px;")
        layout.addWidget(div)

        # Report Type section
        type_lbl = QLabel("REPORT TYPE")
        type_lbl.setObjectName("dlg_section")
        layout.addWidget(type_lbl)

        total  = stats.get("total_students", 0)
        pass_c = stats.get("pass_count", 0)
        fail_c = stats.get("fail_count", 0)

        self.type_group = QButtonGroup(self)
        self.rb_all     = QRadioButton(f"👥  All Students  ({total} students)")
        self.rb_pass    = QRadioButton(f"✅  Passed Only  ({pass_c} students)")
        self.rb_fail    = QRadioButton(f"❌  Failed / ATKT Only  ({fail_c} students)")
        self.rb_top10   = QRadioButton(f"🏆  Top {top_n} Rankers (by SGPA)")
        self.rb_subject = QRadioButton("📊  Subject-wise Performance Report")
        self.rb_all.setChecked(True)
        for rb in [self.rb_all, self.rb_pass, self.rb_fail, self.rb_top10, self.rb_subject]:
            self.type_group.addButton(rb)
            layout.addWidget(rb)

        layout.addSpacing(6)

        # Format section
        fmt_lbl = QLabel("EXPORT FORMAT")
        fmt_lbl.setObjectName("dlg_section")
        layout.addWidget(fmt_lbl)

        self.fmt_group = QButtonGroup(self)
        fmt_row = QHBoxLayout()
        self.rb_excel = QRadioButton("📊  Excel (.xlsx)")
        self.rb_pdf   = QRadioButton("📄  PDF Report")
        self.rb_excel.setChecked(True)
        for rb in [self.rb_excel, self.rb_pdf]:
            self.fmt_group.addButton(rb)
            fmt_row.addWidget(rb)
        fmt_row.addStretch()
        layout.addLayout(fmt_row)

        layout.addStretch()

        # Buttons
        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("secondary")
        cancel_btn.setFixedWidth(110)
        cancel_btn.setCursor(QCursor(Qt.PointingHandCursor))
        cancel_btn.clicked.connect(self.reject)

        generate_btn = QPushButton("📥  Generate Report")
        generate_btn.setFixedWidth(190)
        generate_btn.setCursor(QCursor(Qt.PointingHandCursor))
        generate_btn.clicked.connect(self.accept)

        btn_row.addStretch()
        btn_row.addWidget(cancel_btn)
        btn_row.addSpacing(8)
        btn_row.addWidget(generate_btn)
        layout.addLayout(btn_row)

    def get_selection(self):
        """Returns (report_type, format, top_n)."""
        if self.rb_pass.isChecked():
            rtype = 'pass'
        elif self.rb_fail.isChecked():
            rtype = 'fail'
        elif self.rb_top10.isChecked():
            rtype = 'top10'
        elif self.rb_subject.isChecked():
            rtype = 'subject'
        else:
            rtype = 'all'
        fmt = 'pdf' if self.rb_pdf.isChecked() else 'excel'
        return rtype, fmt, self.top_n


# ==============================================================================
# 8. MAIN APP
# ==============================================================================
class StudentAnalyzerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("VNSGU - Student Result Analyzer | SASCMA")
        self.resize(1380, 870) 
        
        self.db = DatabaseManager()
        self.worker = None
        self.data_loaded = False 
        self.full_df = pd.DataFrame()
        self.is_collapsed = False
        self.current_pdf_path = ""
        
        self.setStyleSheet(VNSGU_STYLESHEET)
        
        self.init_ui()
        self.setup_shortcuts()
        self.show_splash()

    def setup_shortcuts(self):
        # Page navigation shortcuts
        self.shortcut_upload = QShortcut(QKeySequence("Ctrl+U"), self)
        self.shortcut_upload.activated.connect(lambda: self.switch_page(0))  # Home

        # Ctrl+1 to Ctrl+7 — page navigation
        for i, page_idx in enumerate([0, 1, 2, 3, 4, 5, 6], start=1):
            sc = QShortcut(QKeySequence(f"Ctrl+{i}"), self)
            sc.activated.connect(lambda idx=page_idx: self.switch_page(idx))

        # Ctrl+, → Settings (index 7)
        self.shortcut_settings = QShortcut(QKeySequence("Ctrl+,"), self)
        self.shortcut_settings.activated.connect(lambda: self.switch_page(7))

        # Ctrl+8 → College Wise Report (index 8)
        self.shortcut_college = QShortcut(QKeySequence("Ctrl+8"), self)
        self.shortcut_college.activated.connect(lambda: self.switch_page(8))

        # F12 → Settings (legacy)
        self.shortcut_f12 = QShortcut(QKeySequence("F12"), self)
        self.shortcut_f12.activated.connect(lambda: self.switch_page(7))

        # Ctrl+P → Print Report
        self.shortcut_print = QShortcut(QKeySequence("Ctrl+P"), self)
        self.shortcut_print.activated.connect(self.print_report)


    def show_splash(self):
        self.splash = QWidget(None, Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.splash.setObjectName("splash_window")
        self.splash.setFixedSize(640, 400)
        self.splash.setAttribute(Qt.WA_TranslucentBackground)
        self.splash.setStyleSheet("""
            QWidget#splash_window {
                background: transparent;
            }
            QFrame#splash_card {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #0c1a3a, stop:0.5 #1a3a6b, stop:1 #0d9488);
                border-radius: 20px;
                border: 1px solid rgba(255,255,255,0.15);
            }
            QLabel#splash_brand {
                color: white;
                font-size: 26px;
                font-weight: 800;
                letter-spacing: 0.5px;
            }
            QLabel#splash_subtitle {
                color: rgba(255,255,255,0.75);
                font-size: 14px;
                font-weight: 400;
            }
            QLabel#splash_badge {
                color: #f97316;
                font-size: 11px;
                font-weight: 800;
                letter-spacing: 1.5px;
                background-color: rgba(249,115,22,0.12);
                border: 1px solid rgba(249,115,22,0.4);
                padding: 4px 14px;
                border-radius: 4px;
            }
            QProgressBar {
                background: rgba(255,255,255,0.12);
                border: none;
                border-radius: 5px;
                height: 6px;
                text-align: center;
                color: transparent;
            }
            QProgressBar::chunk {
                background: #f97316;
                border-radius: 5px;
            }
        """)

        root = QVBoxLayout(self.splash)
        root.setContentsMargins(16, 16, 16, 16)

        card = QFrame()
        card.setObjectName("splash_card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(48, 38, 48, 38)
        card_layout.setSpacing(10)
        card_layout.setAlignment(Qt.AlignCenter)

        badge = QLabel("VNSGU  ·  SASCMA  ·  STERS")
        badge.setObjectName("splash_badge")
        badge.setAlignment(Qt.AlignCenter)

        logo = QLabel()
        logo.setAlignment(Qt.AlignCenter)
        if os.path.exists(LOGO_PATH):
            pm = QPixmap(LOGO_PATH)
            if not pm.isNull():
                logo.setPixmap(pm.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                logo.setText("🏛️")
        else:
            logo.setText("🏛️")
        logo.setStyleSheet("background: rgba(255,255,255,0.9); border-radius: 50px; padding: 6px; color: #0c1a3a; font-weight: bold; font-size: 40px; min-width:100px; max-width:100px; min-height:100px; max-height:100px;")

        title = QLabel("Student Result Analyzer")
        title.setObjectName("splash_brand")
        title.setAlignment(Qt.AlignCenter)

        subtitle = QLabel("Preparing dashboard · analytics · PDF parser")
        subtitle.setObjectName("splash_subtitle")
        subtitle.setAlignment(Qt.AlignCenter)

        progress = QProgressBar()
        progress.setRange(0, 0)
        progress.setFixedWidth(380)

        card_layout.addWidget(badge, 0, Qt.AlignCenter)
        card_layout.addSpacing(10)
        card_layout.addWidget(logo, 0, Qt.AlignCenter)
        card_layout.addSpacing(6)
        card_layout.addWidget(title)
        card_layout.addWidget(subtitle)
        card_layout.addSpacing(22)
        card_layout.addWidget(progress, 0, Qt.AlignCenter)

        root.addWidget(card)

        screen = QApplication.primaryScreen().availableGeometry()
        self.splash.move(screen.center() - self.splash.rect().center())
        self.splash.show()
        QApplication.processEvents()
        QTimer.singleShot(2200, self.close_splash)

    def close_splash(self):
        self.splash.close()
        self.showMaximized()
        self.switch_page(0)  # Home is now index 0

    def get_subject_headers(self):
        headers = ["Roll No", "Name"]
        for i in range(7):
            name = self.get_subject_name(i)  # Full name, no truncation
            headers.append(name)
        headers.extend(["Total", "SGPA", "Status", "ATKT"])
        return headers

    def get_subject_name(self, index):
        name = "Subject " + str(index + 1)
        if PARSER_LOADED and hasattr(final_result, 'SUBJECTS'):
            if index in final_result.SUBJECTS:
                name = str(final_result.SUBJECTS[index])
        return name

    def get_failed_subjects_for_row(self, row):
        stored = row.get('failed_subjects', '') if hasattr(row, 'get') else ''
        if stored and str(stored).strip() and str(stored).strip() != '-':
            return str(stored)

        failed_subjects = []
        for i in range(7):
            col = f'sub_{i+1}'
            if col not in row:
                continue
            try:
                marks = float(row[col])
            except (TypeError, ValueError):
                continue
            try:
                passing = final_result.EXT_PASSING_MIN[i] if PARSER_LOADED else 18
            except Exception:
                passing = 18
            if marks < passing:
                failed_subjects.append(f"{self.get_subject_name(i)} ({int(marks)})")
        return "\n".join(failed_subjects) if failed_subjects else "-"

    def setup_table(self, table, row_height=38, word_wrap=False):
        table.setAlternatingRowColors(True)
        table.setWordWrap(word_wrap)
        table.setShowGrid(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.verticalHeader().setVisible(False)
        table.verticalHeader().setDefaultSectionSize(row_height)
        table.horizontalHeader().setMinimumSectionSize(58)
        table.horizontalHeader().setStretchLastSection(False)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

    def sort_by_roll_no(self, df):
        if df.empty or 'roll_no' not in df.columns:
            return df
        sorted_df = df.copy()
        sorted_df['_roll_sort'] = pd.to_numeric(sorted_df['roll_no'].astype(str).str.extract(r'(\d+)')[0], errors='coerce')
        sorted_df['_roll_text'] = sorted_df['roll_no'].astype(str)
        sorted_df = sorted_df.sort_values(by=['_roll_sort', '_roll_text'], na_position='last')
        return sorted_df.drop(columns=['_roll_sort', '_roll_text'])

    def sort_by_sgpa_then_roll_no(self, df):
        """Sort order: SGPA (desc) → Total Marks (desc) → Seat/Roll No (asc)"""
        if df.empty or 'sgpa' not in df.columns:
            return df
        sorted_df = df.copy()

        # 1) SGPA — numeric, descending
        sorted_df['_sgpa_sort'] = pd.to_numeric(sorted_df['sgpa'], errors='coerce')

        # 2) Total marks — numeric, descending
        sorted_df['_total_sort'] = pd.to_numeric(
            sorted_df['total'] if 'total' in sorted_df.columns else 0,
            errors='coerce'
        ).fillna(0)

        # 3) Seat / Roll No — strip spaces, extract digits, ascending
        roll_text = (
            sorted_df['roll_no'].astype(str).str.strip()
            if 'roll_no' in sorted_df.columns
            else pd.Series([''] * len(sorted_df), index=sorted_df.index)
        )
        # Try full numeric parse first; fall back to first digit sequence
        roll_numeric = pd.to_numeric(roll_text, errors='coerce')
        roll_from_extract = pd.to_numeric(
            roll_text.str.extract(r'(\d+)')[0], errors='coerce'
        )
        sorted_df['_roll_sort'] = roll_numeric.combine_first(roll_from_extract).fillna(999999)
        sorted_df['_roll_text'] = roll_text

        sorted_df = sorted_df.sort_values(
            by=['_sgpa_sort', '_total_sort', '_roll_sort', '_roll_text'],
            ascending=[False, False, True, True],
            na_position='last',
            kind='mergesort',
        )
        return sorted_df.drop(columns=['_sgpa_sort', '_total_sort', '_roll_sort', '_roll_text'])

    def sgpa_rank_numbers(self, df):
        """
        Dense ranking:
          - Same SGPA → same rank number
          - Next different SGPA → rank + 1 (no gaps ever)
          Example: 1, 2, 2, 3, 3, 3, 3, 4, 4, 4
        """
        ranks = []
        previous_sgpa = object()
        current_rank = 0

        for position, (_, row) in enumerate(df.iterrows(), start=1):
            sgpa = pd.to_numeric(pd.Series([row.get('sgpa')]), errors='coerce').iloc[0]
            if pd.isna(sgpa):
                sgpa = None

            if position == 1 or sgpa != previous_sgpa:
                current_rank = (ranks[-1] + 1) if ranks else 1
                previous_sgpa = sgpa

            ranks.append(current_rank)

        return ranks

    def format_rank_label(self, rank):
        return {1: "🥇 1", 2: "🥈 2", 3: "🥉 3"}.get(rank, str(rank))

    def update_ui_with_subjects(self):
        if hasattr(self, 'combo_subject'):
            self.combo_subject.blockSignals(True) 
            self.combo_subject.clear()
            for i in range(7):
                name = self.get_subject_name(i)
                self.combo_subject.addItem(name)
            self.combo_subject.blockSignals(False)
        
        new_headers = self.get_subject_headers()
        self.table.setColumnCount(len(new_headers))
        self.table.setHorizontalHeaderLabels(new_headers)
        
        # Configure columns widths and resize modes after columns are created/reset
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        self.table.setColumnWidth(0, 80)   # Roll No
        self.table.setColumnWidth(1, 240)  # Name
        for i in range(2, 9):
            self.table.setColumnWidth(i, 130)  # 7 Subjects
        self.table.setColumnWidth(9, 80)   # Total
        self.table.setColumnWidth(10, 80)  # SGPA
        self.table.setColumnWidth(11, 90)  # Status
        self.table.setColumnWidth(12, 80)  # ATKT
        
        self.combo_my_func.blockSignals(True)
        self.combo_my_func.clear()
        for i in range(7):
            name = self.get_subject_name(i)
            self.combo_my_func.addItem(name)
        self.combo_my_func.blockSignals(False)

        # Also update topper combo if it exists
        if hasattr(self, 'combo_topper'):
            self.combo_topper.blockSignals(True)
            self.combo_topper.clear()
            for i in range(7):
                name = self.get_subject_name(i)
                self.combo_topper.addItem(f"Subject {i+1}: {name}")
            self.combo_topper.blockSignals(False)

    def toggle_sidebar(self):
        self.is_collapsed = not self.is_collapsed
        
        expanded_width = 260
        collapsed_width = 68 
        
        self.sidebar_animation = QPropertyAnimation(self.sidebar_container, b"maximumWidth")
        self.sidebar_animation.setDuration(0)   # Instant — removes animation lag
        self.sidebar_animation.setStartValue(self.sidebar_container.width())
        self.sidebar_animation.setEndValue(collapsed_width if self.is_collapsed else expanded_width)
        self.sidebar_animation.setEasingCurve(QEasingCurve.InOutCubic)
        self.sidebar_animation.start()
        
        if self.is_collapsed:
            QTimer.singleShot(140, self.apply_collapsed_state)
        else:
            self.apply_expanded_state()

    def apply_collapsed_state(self):
        self.lbl_college_name.hide()
        self.btn_collapse.hide()
        self.menu_list.hide()
        self.settings_btn.hide()
        self.icon_menu_list.show()
        self.icon_settings_btn.show()
        self.sidebar_header_layout.setAlignment(Qt.AlignCenter)
        self.sidebar_header_layout.setContentsMargins(0, 0, 0, 0)
        self.sidebar_header_layout.setSpacing(10)
        self.logo_lbl.setCursor(QCursor(Qt.PointingHandCursor))
        self.logo_lbl.mousePressEvent = self.toggle_sidebar_event

    def toggle_sidebar_event(self, event):
        self.toggle_sidebar()

    def apply_expanded_state(self):
        self.lbl_college_name.show()
        self.btn_collapse.show()
        self.menu_list.show()
        self.settings_btn.show()
        self.icon_menu_list.hide()
        self.icon_settings_btn.hide()
        self.sidebar_header_layout.setContentsMargins(0, 0, 10, 10)
        self.sidebar_header_layout.setSpacing(0)
        self.sidebar_header_layout.setAlignment(Qt.AlignLeft)
        self.logo_lbl.setCursor(Qt.ArrowCursor)
        self.logo_lbl.mousePressEvent = None

    def init_ui(self):
        cw = QWidget(); self.setCentralWidget(cw)
        main_layout = QHBoxLayout(cw)
        main_layout.setContentsMargins(0,0,0,0)
        main_layout.setSpacing(0)

        # --- 1. SIDEBAR ---
        self.sidebar_container = QFrame()
        self.sidebar_container.setObjectName("sidebar")
        self.sidebar_container.setFixedWidth(260)
        self.sidebar_container.setMinimumWidth(68)
        self.sidebar_container.setMaximumWidth(260)
        
        sidebar_layout = QVBoxLayout(self.sidebar_container)
        sidebar_layout.setContentsMargins(0,0,0,0)
        sidebar_layout.setSpacing(0)
        sidebar_layout.setAlignment(Qt.AlignTop)   # Fix: items from top, no empty gap
        
        # Sidebar Header with Logo
        sidebar_header = QFrame()
        sidebar_header.setObjectName("sidebar_logo_container")
        sidebar_header.setFixedHeight(70)
        self.sidebar_header_layout = QHBoxLayout(sidebar_header)
        self.sidebar_header_layout.setContentsMargins(10, 8, 10, 8)
        self.sidebar_header_layout.setSpacing(8)
        
        self.logo_lbl = QLabel()
        self.logo_lbl.setObjectName("sidebar_logo")
        self.logo_lbl.setFixedSize(46, 46)
        if os.path.exists(LOGO_PATH):
            pm = QPixmap(LOGO_PATH)
            if not pm.isNull():
                self.logo_lbl.setPixmap(pm.scaled(44, 44, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                self.logo_lbl.setText("🏛️")
                self.logo_lbl.setStyleSheet("font-size: 22px; border: none;")
        else:
            self.logo_lbl.setText("🏛️")
            self.logo_lbl.setStyleSheet("font-size: 22px; border: none;")
        self.logo_lbl.setAlignment(Qt.AlignCenter)

        self.lbl_college_name = QLabel("SASCMA - STERS")
        self.lbl_college_name.setObjectName("sidebar_college_name")
        self.lbl_college_name.setStyleSheet("padding: 0px; background: transparent;")

        self.btn_collapse = QPushButton("◀")
        self.btn_collapse.setObjectName("toggle_btn")
        self.btn_collapse.setFixedSize(28, 28)
        self.btn_collapse.clicked.connect(self.toggle_sidebar)
        
        self.sidebar_header_layout.addWidget(self.logo_lbl)
        self.sidebar_header_layout.addWidget(self.lbl_college_name)
        self.sidebar_header_layout.addStretch()
        self.sidebar_header_layout.addWidget(self.btn_collapse)

        sidebar_layout.addWidget(sidebar_header)

        # Stack indices: 0=Upload, 1=Dashboard, 2=TopN, 3=Failed, 4=Subject, 5=SGPA, 6=Lookup, 7=Settings, 8=CollegeWise
        menu_items_data = [
            ("📤  Upload PDF",        "UPLOAD PDF              [Ctrl+1]",  0),
            ("📊  Dashboard",         "DASHBOARD & ANALYTICS   [Ctrl+2]",  1),
            ("🏆  Top N Rankers",    "TOP N RANKERS           [Ctrl+3]",  2),
            ("❌  Failed Students",   "ATKT / FAIL LIST        [Ctrl+4]",  3),
            ("📈  Subject Analysis",  "SUBJECT WISE REPORT     [Ctrl+5]",  4),
            ("📉  SGPA Analysis",     "SGPA & GRADE ANALYTICS  [Ctrl+6]",  5),
            ("🔍  Student Lookup",    "SEARCH STUDENT          [Ctrl+7]",  6),
            ("🏫  College Wise",      "COLLEGE WISE REPORT     [Ctrl+8]",  8),
        ]
        icon_items_data = [
            ("📤", 0), ("📊", 1), ("🏆", 2), ("❌", 3),
            ("📈", 4), ("📉", 5), ("🔍", 6), ("🏫", 8)
        ]

        # Full text menu list (visible when expanded)
        self.menu_list = QListWidget()
        self.menu_list.setObjectName("menu_list")
        self.menu_list.currentRowChanged.connect(self.on_sidebar_click)
        self.menu_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.menu_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.menu_list.setSizeAdjustPolicy(QListWidget.AdjustToContents)

        for text, tooltip, idx in menu_items_data:
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, idx)
            item.setToolTip(tooltip)
            self.menu_list.addItem(item)

        # Icon-only list (visible when collapsed)
        self.icon_menu_list = QListWidget()
        self.icon_menu_list.setObjectName("icon_menu_list")
        self.icon_menu_list.currentRowChanged.connect(self.on_icon_sidebar_click)
        self.icon_menu_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.icon_menu_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.icon_menu_list.setSizeAdjustPolicy(QListWidget.AdjustToContents)
        self.icon_menu_list.hide()  # hidden by default

        for icon, idx in icon_items_data:
            item = QListWidgetItem(icon)
            item.setData(Qt.UserRole, idx)
            item.setTextAlignment(Qt.AlignCenter)
            self.icon_menu_list.addItem(item)

        sidebar_layout.addWidget(self.menu_list)
        sidebar_layout.addWidget(self.icon_menu_list)

        # Divider
        div2 = QFrame()
        div2.setFixedHeight(1)
        div2.setStyleSheet("background-color: rgba(255,255,255,0.08); margin: 0px;")
        sidebar_layout.addWidget(div2)

        self.settings_btn = QPushButton("⚙  Settings")
        self.settings_btn.setObjectName("settings_btn")
        self.settings_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.settings_btn.clicked.connect(lambda: self.switch_page(7))
        sidebar_layout.addWidget(self.settings_btn)

        # Icon-only settings button (collapsed state)
        self.icon_settings_btn = QPushButton("⚙")
        self.icon_settings_btn.setObjectName("icon_settings_btn")
        self.icon_settings_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.icon_settings_btn.clicked.connect(lambda: self.switch_page(7))
        self.icon_settings_btn.hide()
        sidebar_layout.addWidget(self.icon_settings_btn)
        
        main_layout.addWidget(self.sidebar_container)
        
        # --- 2. MAIN CONTENT ---
        content_container = QWidget()
        content_layout = QVBoxLayout(content_container)
        content_layout.setContentsMargins(0,0,0,0)
        content_layout.setSpacing(0)

        # Top Header — Teal Banner (VNSGU style)
        top_header = QFrame()
        top_header.setObjectName("main_header")
        th_layout = QHBoxLayout(top_header)
        th_layout.setContentsMargins(20, 0, 20, 0)
        
        # Hamburger + Title
        left_row = QHBoxLayout()
        left_row.setSpacing(14)
        
        title_lbl = QLabel("Student Result Analyzer")
        title_lbl.setObjectName("main_title")
        left_row.addWidget(title_lbl)
        
        th_layout.addLayout(left_row)
        th_layout.addStretch()
        
        # Academic year label + date
        self.acad_lbl = QLabel("Academic Year : ----")
        self.acad_lbl.setStyleSheet("color: rgba(255,255,255,0.8); font-size: 12px; font-weight: 600; background: rgba(255,255,255,0.12); border-radius: 4px; padding: 4px 12px;")
        th_layout.addWidget(self.acad_lbl)
        th_layout.addSpacing(12)
        
        date_lbl = QLabel(QDate.currentDate().toString("dddd, dd MMM yyyy"))
        date_lbl.setObjectName("header_date")
        th_layout.addWidget(date_lbl)
        th_layout.addSpacing(10)
        
        welcome_lbl = QLabel("Welcome, Admin")
        welcome_lbl.setObjectName("header_welcome")
        th_layout.addWidget(welcome_lbl)

        content_layout.addWidget(top_header)

        # Stack — 0=Home, 1=Dashboard, 2=TopN, 3=Failed, 4=Subject, 5=SGPA, 6=Lookup, 7=Settings
        self.stack = QStackedWidget()
        
        self.upload_page = self.create_upload_page()            # index 0 = Home
        self.stack.addWidget(self.upload_page)

        self.dashboard_page = self.create_dashboard_page()      # index 1 = Dashboard
        self.stack.addWidget(self.dashboard_page)
        
        self.top_n_page = self.create_top_n_page()              # index 2
        self.stack.addWidget(self.top_n_page)
        
        self.failed_page = self.create_failed_page()            # index 3
        self.stack.addWidget(self.failed_page)
        
        self.my_functions_page = self.create_my_functions_page()  # index 4
        self.stack.addWidget(self.my_functions_page)

        self.sgpa_page = self.create_sgpa_analysis_page()       # index 5 — NEW
        self.stack.addWidget(self.sgpa_page)

        self.lookup_page = self.create_student_lookup_page()    # index 6 — NEW
        self.stack.addWidget(self.lookup_page)

        self.settings_page = self.create_settings_page()        # index 7 (was 5)
        self.stack.addWidget(self.settings_page)

        self.college_wise_page = self.create_college_wise_page()  # index 8
        self.stack.addWidget(self.college_wise_page)

        content_layout.addWidget(self.stack, 1) 
        
        # Footer — Dark Navy
        footer = QFrame()
        footer.setObjectName("ifooter")
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(20, 0, 20, 0)
        footer_lbl = QLabel("© 2026 VNSGU · SASCMA - STERS Student Result Analyzer. All Rights Reserved.")
        footer_lbl.setObjectName("footer_lbl")
        footer_layout.addWidget(footer_lbl)
        footer_layout.addStretch()
        ver_lbl = QLabel("v2.1  |  Ctrl+1–7 = Pages  |  Ctrl+, = Settings  |  Ctrl+P = Print")
        ver_lbl.setObjectName("footer_lbl")
        footer_layout.addWidget(ver_lbl)
        content_layout.addWidget(footer)

        main_layout.addWidget(content_container)

        # Initial Lock — only Home (index 0) enabled at start
        for i in range(self.menu_list.count()):
            item = self.menu_list.item(i)
            if item.data(Qt.UserRole) != 0:
                item.setFlags(Qt.NoItemFlags)
        for i in range(self.icon_menu_list.count()):
            item = self.icon_menu_list.item(i)
            if item.data(Qt.UserRole) != 0:
                item.setFlags(Qt.NoItemFlags)

    def on_sidebar_click(self, row):
        item = self.menu_list.item(row)
        if item and item.flags() & Qt.ItemIsEnabled:
            idx = item.data(Qt.UserRole)
            self.switch_page(idx)

    def on_icon_sidebar_click(self, row):
        item = self.icon_menu_list.item(row)
        if item and item.flags() & Qt.ItemIsEnabled:
            idx = item.data(Qt.UserRole)
            self.switch_page(idx)

    def switch_page(self, index):
        # Sync both menu lists
        for i in range(self.menu_list.count()):
            item = self.menu_list.item(i)
            if item and item.data(Qt.UserRole) == index:
                self.menu_list.setCurrentItem(item)
                break
        for i in range(self.icon_menu_list.count()):
            item = self.icon_menu_list.item(i)
            if item and item.data(Qt.UserRole) == index:
                self.icon_menu_list.setCurrentItem(item)
                break
        
        self.stack.setCurrentIndex(index)
        # Indices: 0=Home, 1=Dashboard, 2=TopN, 3=Failed, 4=Subject, 5=SGPA, 6=Lookup, 7=Settings
        if index == 1: self.refresh_dashboard()
        if index == 2: self.update_top_n_analysis()
        if index == 3: self.refresh_failed_page()
        if index == 4: self.refresh_my_functions()
        if index == 5: self.refresh_sgpa_analysis()
        if index == 6: pass  # Student Lookup — user types their own query
        if index == 7: self.refresh_settings_page()
        if index == 8: self.refresh_college_wise()

    def unlock_features(self):
        self.data_loaded = True
        for i in range(self.menu_list.count()):
            item = self.menu_list.item(i)
            if item:
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsUserCheckable)
        for i in range(self.icon_menu_list.count()):
            item = self.icon_menu_list.item(i)
            if item:
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsUserCheckable)
        
        # Update academic year from extracted PDF data
        try:
            ay = getattr(final_result, 'ACADEMIC_YEAR', None) or ""
            if ay:
                self.acad_lbl.setText(f"Academic Year : {ay}")
        except Exception:
            pass

        self.switch_page(1)  # Go to Dashboard after upload
        self.update_ui_with_subjects() 
        print("Features Unlocked & UI Updated.")

    # --- Pages ---

    def create_dashboard_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        # Page Title Row
        page_title_row = QHBoxLayout()
        page_title = QLabel("Dashboard")
        page_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0c1a3a;")
        page_title_row.addWidget(page_title)
        page_title_row.addStretch()
        
        # Search + Export
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("🔍  Search by Name or Roll No...")
        self.search_bar.setFixedWidth(280)
        self.search_bar.textChanged.connect(self.filter_table)
        page_title_row.addWidget(self.search_bar)
        page_title_row.addSpacing(8)
        
        btn_export = QPushButton("⬇  Export Excel")
        btn_export.setCursor(QCursor(Qt.PointingHandCursor))
        btn_export.clicked.connect(lambda: self.export_to_excel(self.db.get_all_data()))
        page_title_row.addWidget(btn_export)

        btn_export_pdf = QPushButton("📄  Export PDF")
        btn_export_pdf.setObjectName("secondary")
        btn_export_pdf.setCursor(QCursor(Qt.PointingHandCursor))
        btn_export_pdf.clicked.connect(lambda: self.export_to_pdf(self.db.get_all_data()))
        page_title_row.addWidget(btn_export_pdf)
        
        layout.addLayout(page_title_row)
        
        # KPI Cards Row
        kpi_row = QHBoxLayout()
        kpi_row.setSpacing(14)
        self.kpi_total = KPICard("Total Students", "0", "#0c1a3a", "👥")
        self.kpi_pass  = KPICard("Pass %",          "0%", "#0d9488", "✅")
        self.kpi_fail  = KPICard("Fail %",           "0%", "#dc2626", "❌")
        self.kpi_topper= KPICard("Top Scorer",       "N/A","#f97316", "🏆")
        
        for card in [self.kpi_total, self.kpi_pass, self.kpi_fail, self.kpi_topper]:
            kpi_row.addWidget(card)
        layout.addLayout(kpi_row)
        
        # Charts Section
        charts_outer, charts_content, _, charts_inner = make_section_card("📈  Analytics Overview", collapsible=False)
        charts_row = QHBoxLayout()
        charts_row.setSpacing(14)
        
        c1 = QFrame(); c1.setObjectName("card"); l1 = QVBoxLayout(c1)
        lbl_pie = QLabel("Pass / Fail Ratio")
        lbl_pie.setStyleSheet("font-weight:700; color:#0c1a3a; font-size:13px;")
        self.pie_canvas = MplCanvas(self, width=5, height=3.5) 
        l1.addWidget(lbl_pie); l1.addWidget(self.pie_canvas)
        
        c2 = QFrame(); c2.setObjectName("card"); l2 = QVBoxLayout(c2)
        lbl_bar = QLabel("Grade Distribution")
        lbl_bar.setStyleSheet("font-weight:700; color:#0c1a3a; font-size:13px;")
        self.bar_canvas = MplCanvas(self, width=5, height=3.5) 
        l2.addWidget(lbl_bar); l2.addWidget(self.bar_canvas)
        c1.setMinimumHeight(350)
        c2.setMinimumHeight(350)
        
        charts_row.addWidget(c1); charts_row.addWidget(c2)
        charts_inner.addLayout(charts_row)
        layout.addWidget(charts_outer)
        
        # Student Table Section
        table_outer, table_content, _, table_inner = make_section_card("👨‍🎓  Student Results Table", collapsible=False)
        
        self.table = QTableWidget()
        self.table.setColumnCount(12) 
        self.table.setHorizontalHeaderLabels(self.get_subject_headers())
        self.setup_table(self.table, row_height=36)
        self.table.setMinimumHeight(420) 
        self.table.setSortingEnabled(False)
        
        header = self.table.horizontalHeader()
        # Roll No: fixed width
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        # Name: stretch to fill available space
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        # Subject columns (2-8): Interactive so full name shows, user can resize
        for i in range(2, 9):
            header.setSectionResizeMode(i, QHeaderView.Interactive)
            self.table.setColumnWidth(i, 130)  # default 130px wide (fits full subject name)
        # Total, SGPA, Status, ATKT
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(10, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(11, QHeaderView.ResizeToContents)
        # Allow horizontal scroll so all columns are accessible
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        header.setMinimumSectionSize(80)
        
        table_inner.addWidget(self.table)
        layout.addWidget(table_outer)
        
        layout.addStretch()
        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def create_upload_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setAlignment(Qt.AlignCenter)
        
        # Page title
        upload_title = QLabel("📤  Upload PDF")
        upload_title.setStyleSheet("font-size: 24px; font-weight: 800; color: #0c1a3a; margin-bottom: 4px;")
        upload_title.setAlignment(Qt.AlignCenter)
        layout.addWidget(upload_title)
        
        upload_sub = QLabel("Upload the VNSGU examination result PDF to begin analysis")
        upload_sub.setStyleSheet("color: #64748b; font-size: 13px; margin-bottom: 24px;")
        upload_sub.setAlignment(Qt.AlignCenter)
        layout.addWidget(upload_sub)
        
        drop_zone = QFrame()
        drop_zone.setObjectName("drop_zone")
        drop_zone.setFixedSize(520, 320)
        dz_layout = QVBoxLayout(drop_zone)
        dz_layout.setAlignment(Qt.AlignCenter)
        dz_layout.setSpacing(12)
        
        icon = QLabel("📄")
        icon.setStyleSheet("font-size: 64px; color: #0d9488; background: transparent;")
        icon.setAlignment(Qt.AlignCenter)
        
        title = QLabel("Drag & Drop PDF File Here")
        title.setStyleSheet("font-size: 18px; font-weight: 700; color: #0c1a3a; background: transparent; margin-top: 6px;")
        title.setAlignment(Qt.AlignCenter)
        
        desc = QLabel("or click the button below to browse files")
        desc.setStyleSheet("color: #64748b; font-size: 13px; background: transparent;")
        desc.setAlignment(Qt.AlignCenter)
        
        divider_lbl = QLabel("—————  OR  —————")
        divider_lbl.setStyleSheet("color: #cbd5e1; font-size: 12px; background: transparent;")
        divider_lbl.setAlignment(Qt.AlignCenter)
        
        self.browse_btn = QPushButton("📂  Browse PDF File")
        self.browse_btn.setFixedSize(220, 46)
        self.browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #0c1a3a;
                border: none;
                border-radius: 8px;
                color: white;
                font-size: 14px;
                font-weight: 700;
            }
            QPushButton:hover { background-color: #1a3a6b; }
        """)
        self.browse_btn.setCursor(QCursor(Qt.PointingHandCursor))
        
        dz_layout.addWidget(icon)
        dz_layout.addWidget(title)
        dz_layout.addWidget(desc)
        dz_layout.addSpacing(8)
        dz_layout.addWidget(divider_lbl)
        dz_layout.addSpacing(8)
        dz_layout.addWidget(self.browse_btn, 0, Qt.AlignCenter)
        
        layout.addWidget(drop_zone, 0, Qt.AlignCenter)
        
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setFixedWidth(520)
        self.status_lbl = QLabel("Ready to upload")
        self.status_lbl.setAlignment(Qt.AlignCenter)
        self.status_lbl.setStyleSheet("margin-top: 12px; color: #64748b; font-size: 13px;")
        
        layout.addSpacing(16)
        layout.addWidget(self.progress, 0, Qt.AlignCenter)
        layout.addWidget(self.status_lbl)
        
        self.browse_btn.clicked.connect(self.browse_pdf)
        page.setAcceptDrops(True)
        page.dragEnterEvent = self.drag_enter_event
        page.dropEvent = self.drop_event

        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def create_top_n_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        # Title
        pg_title = QLabel("🏆  Top N Rankers")
        pg_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0c1a3a;")
        layout.addWidget(pg_title)
        
        # Controls Card
        ctrl_card = QFrame()
        ctrl_card.setObjectName("card")
        ctrl_card.setFixedHeight(70)
        ctrl_layout = QHBoxLayout(ctrl_card)
        ctrl_layout.setContentsMargins(18, 10, 18, 10)
        
        ctrl_layout.addWidget(QLabel("Show Top N Ranks:"))
        
        self.spin_n = QSpinBox()
        self.spin_n.setRange(1, 100)
        self.spin_n.setValue(10)
        self.spin_n.setFixedWidth(100)
        self.spin_n.setFixedHeight(36)
        
        students_lbl = QLabel("ranks by SGPA  (same SGPA = same rank, no gaps)")
        students_lbl.setStyleSheet("color: #64748b;")
        
        self.btn_show_top_n = QPushButton("🔍  Analyze")
        self.btn_show_top_n.setFixedHeight(36)
        self.btn_show_top_n.clicked.connect(self.update_top_n_analysis)
        self.btn_show_top_n.setCursor(QCursor(Qt.PointingHandCursor))
        
        ctrl_layout.addWidget(self.spin_n)
        ctrl_layout.addWidget(students_lbl)
        ctrl_layout.addSpacing(20)
        ctrl_layout.addWidget(self.btn_show_top_n)
        ctrl_layout.addStretch()
        layout.addWidget(ctrl_card)
        
        # Table Section
        table_outer, _, _, table_inner = make_section_card("🏆  Ranking Table", collapsible=False)
        self.top_n_table = QTableWidget()
        self.top_n_table.setColumnCount(6)
        self.top_n_table.setHorizontalHeaderLabels(["Rank", "Roll No", "Name", "SGPA", "Total Marks", "Status"])
        self.setup_table(self.top_n_table, row_height=36)
        self.top_n_table.setMinimumHeight(400)
        top_header = self.top_n_table.horizontalHeader()
        top_header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        top_header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        top_header.setSectionResizeMode(2, QHeaderView.Stretch)
        top_header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        top_header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        top_header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        table_inner.addWidget(self.top_n_table)
        layout.addWidget(table_outer)
        layout.addStretch()

        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def update_top_n_analysis(self):
        n = self.spin_n.value()
        df = self.db.get_all_data()
        if df.empty:
            QMessageBox.warning(self, "No Data", "Please upload a PDF first.")
            return

        # Sort entire dataset by SGPA desc, roll_no asc
        try:
            sorted_df = self.sort_by_sgpa_then_roll_no(df).reset_index(drop=True)
        except Exception:
            sorted_df = df.reset_index(drop=True)

        # Compute dense ranks for ALL students
        all_ranks = self.sgpa_rank_numbers(sorted_df)

        # Keep only students whose dense rank <= N
        # This ensures rank goes from 1 up to N (or max rank if fewer unique SGPA groups)
        filtered = [(r_data, rank_no)
                    for (_, r_data), rank_no in zip(sorted_df.iterrows(), all_ranks)
                    if rank_no <= n]

        self.top_n_table.setRowCount(len(filtered))
        for r_idx, (r_data, rank_no) in enumerate(filtered):
            rank_text = self.format_rank_label(rank_no)
            self.top_n_table.setItem(r_idx, 0, QTableWidgetItem(rank_text))
            self.top_n_table.setItem(r_idx, 1, QTableWidgetItem(str(r_data['roll_no'])))
            self.top_n_table.setItem(r_idx, 2, QTableWidgetItem(str(r_data['name'])))
            self.top_n_table.setItem(r_idx, 3, QTableWidgetItem(str(r_data['sgpa'])))
            self.top_n_table.setItem(r_idx, 4, QTableWidgetItem(str(r_data['total'])))

            status_item = QTableWidgetItem(str(r_data['status']))
            if r_data['status'] == 'FAIL':
                status_item.setForeground(QColor("#dc2626"))
                status_item.setBackground(QColor("#fef2f2"))
            else:
                status_item.setForeground(QColor("#16a34a"))
                status_item.setBackground(QColor("#f0fdf4"))
            self.top_n_table.setItem(r_idx, 5, status_item)

    def create_failed_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        # Title Row
        header_row = QHBoxLayout()
        pg_title = QLabel("❌  Failed Students / ATKT Analysis")
        pg_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0c1a3a;")
        header_row.addWidget(pg_title)
        header_row.addStretch()
        btn_export_fail = QPushButton("⬇  Export Failed List")
        btn_export_fail.setCursor(QCursor(Qt.PointingHandCursor))
        btn_export_fail.clicked.connect(self.export_failed_students)
        header_row.addWidget(btn_export_fail)
        layout.addLayout(header_row)

        table_outer, _, _, table_inner = make_section_card("❌  Failed Students Detail", collapsible=False)
        self.failed_table = QTableWidget()
        self.failed_table.setColumnCount(6)
        self.failed_table.setHorizontalHeaderLabels(["Roll No", "Name", "Total Marks", "ATKTs", "Failed Subjects", "Status"])
        self.setup_table(self.failed_table, row_height=68, word_wrap=True)
        self.failed_table.setMinimumHeight(400)
        self.failed_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.failed_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.failed_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.failed_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.failed_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.failed_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        table_inner.addWidget(self.failed_table)
        layout.addWidget(table_outer)
        layout.addStretch()
        
        if self.data_loaded:
            self.refresh_failed_page()

        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def refresh_failed_page(self):
        df = self.db.get_all_data()
        failed_df = df[df['status'] == 'FAIL']
        
        self.failed_table.setRowCount(len(failed_df))
        for r_idx, (_, r_data) in enumerate(failed_df.iterrows()):
            self.failed_table.setItem(r_idx, 0, QTableWidgetItem(str(r_data['roll_no'])))
            self.failed_table.setItem(r_idx, 1, QTableWidgetItem(str(r_data['name'])))
            self.failed_table.setItem(r_idx, 2, QTableWidgetItem(str(r_data['total'])))
            atkt_item = QTableWidgetItem(str(r_data['atkt_count']))
            atkt_item.setBackground(QColor("#fef2f2"))
            atkt_item.setForeground(QColor("#dc2626"))
            self.failed_table.setItem(r_idx, 3, atkt_item)
            self.failed_table.setItem(r_idx, 4, QTableWidgetItem(self.get_failed_subjects_for_row(r_data)))
            
            status_item = QTableWidgetItem("FAIL")
            status_item.setForeground(QColor("#ffffff"))
            status_item.setBackground(QColor("#dc2626"))
            self.failed_table.setItem(r_idx, 5, status_item)
            line_count = max(1, self.get_failed_subjects_for_row(r_data).count("\n") + 1)
            self.failed_table.setRowHeight(r_idx, max(68, 34 * line_count))

    def export_failed_students(self):
        df = self.db.get_all_data()
        if df.empty:
            QMessageBox.warning(self, "No Data", "No data available to export.")
            return
        failed_df = df[df['status'] == 'FAIL'].copy()
        if failed_df.empty:
            QMessageBox.information(self, "No Failed Students", "No failed students found.")
            return
        failed_df['failed_subjects'] = failed_df.apply(self.get_failed_subjects_for_row, axis=1)
        self.export_to_excel(failed_df, "failed_students.xlsx")

    def create_my_functions_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        pg_title = QLabel("📈  Subject Wise Analysis")
        pg_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0c1a3a;")
        layout.addWidget(pg_title)
        
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.currentChanged.connect(self.refresh_my_functions)
        
        # Tab 1: Subject Analysis
        self.tab_subject = QWidget()
        sub_layout = QVBoxLayout(self.tab_subject)
        sub_layout.setContentsMargins(16, 16, 16, 16)
        sub_layout.setSpacing(12)
        
        ctrl_row = QHBoxLayout()
        ctrl_row.addWidget(QLabel("Select Subject:"))
        ctrl_row.addSpacing(8)
        
        self.combo_my_func = QComboBox()
        self.combo_my_func.setFixedWidth(260)
        self.combo_my_func.currentIndexChanged.connect(self.update_my_functions_charts)
        ctrl_row.addWidget(self.combo_my_func)
        ctrl_row.addStretch()
        sub_layout.addLayout(ctrl_row)
        
        content_layout = QHBoxLayout()
        content_layout.setSpacing(14)
        
        # Left: Stats Card
        stats_card = QFrame()
        stats_card.setObjectName("card")
        stats_card.setFixedWidth(220)
        sl = QVBoxLayout(stats_card)
        sl.setSpacing(10)
        
        stats_title = QLabel("Subject Stats")
        stats_title.setStyleSheet("font-size: 14px; font-weight: 700; color: #0c1a3a; margin-bottom: 4px;")
        sl.addWidget(stats_title)
        
        def stat_row(label, attr_name, color="#0c1a3a"):
            row = QFrame()
            row.setStyleSheet("background: #f8fafc; border-radius: 6px; padding: 6px;")
            rl = QVBoxLayout(row)
            rl.setContentsMargins(10, 6, 10, 6)
            rl.setSpacing(2)
            lbl = QLabel(label)
            lbl.setStyleSheet("color: #64748b; font-size: 11px; font-weight: 600;")
            val = QLabel("—")
            val.setStyleSheet(f"color: {color}; font-size: 18px; font-weight: 800;")
            rl.addWidget(lbl)
            rl.addWidget(val)
            sl.addWidget(row)
            setattr(self, attr_name, val)
        
        stat_row("Total Students", "my_ana_total", "#0c1a3a")
        stat_row("Passed",         "my_ana_pass",  "#0d9488")
        stat_row("Failed",         "my_ana_fail",  "#dc2626")
        stat_row("Average Marks",  "my_ana_avg",   "#f97316")
        stat_row("Highest Marks",  "my_ana_high",  "#16a34a")
        stat_row("Lowest Marks",   "my_ana_low",   "#64748b")
        sl.addStretch()
        
        # Right: Chart Card
        chart_card = QFrame()
        chart_card.setObjectName("card")
        cl = QVBoxLayout(chart_card)
        self.my_ana_chart_title = QLabel("Subject Performance Overview")
        self.my_ana_chart_title.setStyleSheet("font-weight:700; color:#0c1a3a; font-size:13px;")
        self.my_ana_canvas = MplCanvas(self, width=6, height=4.5)
        cl.addWidget(self.my_ana_chart_title)
        cl.addWidget(self.my_ana_canvas)
        
        content_layout.addWidget(stats_card)
        content_layout.addWidget(chart_card)
        sub_layout.addLayout(content_layout)
        
        self.tabs.addTab(self.tab_subject, "📊  Subject Wise Analysis")
        
        # Tab 2: ATKT Calculator
        self.tab_atkt = QWidget()
        atkt_layout = QVBoxLayout(self.tab_atkt)
        atkt_layout.setContentsMargins(16, 16, 16, 16)
        atkt_layout.setSpacing(12)
        
        atkt_header = QHBoxLayout()
        atkt_header.addWidget(QLabel("ATKT Calculation per Subject"))
        atkt_header.addStretch()
        btn_recalculate = QPushButton("🔄  Recalculate")
        btn_recalculate.setCursor(QCursor(Qt.PointingHandCursor))
        btn_recalculate.clicked.connect(self.calculate_atkt)
        atkt_header.addWidget(btn_recalculate)
        atkt_layout.addLayout(atkt_header)
        
        self.atkt_table = QTableWidget()
        self.atkt_table.setColumnCount(5)
        self.atkt_table.setHorizontalHeaderLabels(["Subject Name", "Total Students", "Total Failed", "Total ATKTs", "Failure Rate (%)"])
        self.setup_table(self.atkt_table, row_height=36)
        atkt_layout.addWidget(self.atkt_table)
        
        self.tabs.addTab(self.tab_atkt, "📋  ATKT Calculator")

        # Tab 3: Subject Topper
        self.tab_topper = QWidget()
        topper_layout = QVBoxLayout(self.tab_topper)
        topper_layout.setContentsMargins(16, 16, 16, 16)
        topper_layout.setSpacing(12)

        topper_header = QHBoxLayout()
        topper_lbl = QLabel("Select Subject to View Topper:")
        topper_lbl.setStyleSheet("font-weight: 600; color: #0c1a3a;")
        topper_header.addWidget(topper_lbl)
        topper_header.addSpacing(8)

        self.combo_topper = QComboBox()
        self.combo_topper.setFixedWidth(280)
        topper_header.addWidget(self.combo_topper)

        topper_all_btn = QPushButton("🏅  Show All Toppers")
        topper_all_btn.setCursor(QCursor(Qt.PointingHandCursor))
        topper_all_btn.clicked.connect(lambda: self.refresh_subject_topper(show_all=True))
        topper_header.addWidget(topper_all_btn)

        topper_search_btn = QPushButton("🔍  Find Topper")
        topper_search_btn.setObjectName("secondary")
        topper_search_btn.setCursor(QCursor(Qt.PointingHandCursor))
        topper_search_btn.clicked.connect(lambda: self.refresh_subject_topper(show_all=False))
        topper_header.addWidget(topper_search_btn)

        topper_header.addStretch()
        topper_layout.addLayout(topper_header)

        self.topper_table = QTableWidget()
        self.topper_table.setColumnCount(7)
        self.topper_table.setHorizontalHeaderLabels([
            "Rank", "Roll No", "Student Name", "Subject", "Marks", "SGPA", "Status"
        ])
        self.setup_table(self.topper_table, row_height=36)
        th = self.topper_table.horizontalHeader()
        th.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        th.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        th.setSectionResizeMode(2, QHeaderView.Stretch)
        th.setSectionResizeMode(3, QHeaderView.Stretch)
        th.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        th.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        th.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        topper_layout.addWidget(self.topper_table)

        self.tabs.addTab(self.tab_topper, "🏆  Subject Topper")
        layout.addWidget(self.tabs)

        if self.data_loaded:
            self.update_my_functions_charts(0)

        layout.addStretch()
        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def update_my_functions_charts(self, index=None):
        current_tab = self.tabs.currentIndex() if hasattr(self, 'tabs') else 0
        
        if current_tab == 0:
            subject_idx = self.combo_my_func.currentIndex()
            if subject_idx < 0:
                subject_idx = 0
            
            df = self.db.get_all_data()
            if df.empty: return

            subject_col = f'sub_{subject_idx + 1}'
            if subject_col not in df.columns:
                return
            marks = df[subject_col]
            valid_marks = marks[marks > 0]
            
            if len(valid_marks) == 0:
                for attr in ['my_ana_total','my_ana_pass','my_ana_fail','my_ana_avg','my_ana_high','my_ana_low']:
                    getattr(self, attr).setText("0")
                self.my_ana_canvas.axes.clear()
                self.my_ana_canvas.draw()
                return

            total_studs = len(df)
            try:
                passing = final_result.EXT_PASSING_MIN[subject_idx] if PARSER_LOADED else 18
            except Exception:
                passing = 18
            subject_name = self.get_subject_name(subject_idx)
            if 'failed_subjects' in df.columns and df['failed_subjects'].astype(str).str.strip().ne('').any():
                failed_mask = df['failed_subjects'].astype(str).str.contains(subject_name, regex=False, na=False)
                failed_studs = df[failed_mask]
            else:
                failed_studs = df[(df[subject_col] < passing)]
            fail_count = len(failed_studs)
            pass_count = total_studs - fail_count
            
            self.my_ana_total.setText(str(total_studs))
            self.my_ana_pass.setText(str(pass_count))
            self.my_ana_fail.setText(str(fail_count))
            self.my_ana_avg.setText(f"{valid_marks.mean():.2f}")
            self.my_ana_high.setText(str(int(valid_marks.max())))
            self.my_ana_low.setText(str(int(valid_marks.min())))
            self.my_ana_chart_title.setText(f"{subject_name} — Marks Distribution")
            
            palette = ['#0d9488', '#0891b2', '#16a34a', '#ca8a04', '#ea580c', '#7c3aed', '#dc2626']
            chart_color = palette[subject_idx % len(palette)]
            self.my_ana_canvas.axes.clear()
            self.my_ana_canvas.axes.set_facecolor("#f8fafc")
            self.my_ana_canvas.fig.patch.set_facecolor('#ffffff')
            self.my_ana_canvas.axes.hist(valid_marks, bins=10, color=chart_color, edgecolor='#ffffff', linewidth=1.2, alpha=0.88)
            self.my_ana_canvas.axes.axvline(passing, color='#dc2626', linestyle='--', linewidth=2, label=f'Pass Mark: {passing}')
            self.my_ana_canvas.axes.set_title(f"{subject_name} Marks Distribution", fontsize=12, fontweight='bold', color='#0c1a3a')
            self.my_ana_canvas.axes.set_xlabel("Marks", color='#64748b')
            self.my_ana_canvas.axes.set_ylabel("No. of Students", color='#64748b')
            self.my_ana_canvas.axes.grid(axis='y', color='#e2e8f0', linewidth=0.8)
            self.my_ana_canvas.axes.legend(frameon=False, fontsize=10)
            self.my_ana_canvas.axes.spines['top'].set_visible(False)
            self.my_ana_canvas.axes.spines['right'].set_visible(False)
            self.my_ana_canvas.axes.tick_params(colors='#64748b')
            self.my_ana_canvas.draw()
            
        elif current_tab == 1:
            self.calculate_atkt()

    def calculate_atkt(self):
        df = self.db.get_all_data()
        if df.empty:
            return

        self.atkt_table.setRowCount(7)
        
        for i in range(7):
            if f'sub_{i+1}' in df.columns:
                col = f'sub_{i+1}'
                total = len(df)
                try:
                    passing = final_result.EXT_PASSING_MIN[i] if PARSER_LOADED else 18
                except Exception:
                    passing = 18
                subject_name = self.get_subject_name(i)
                if 'failed_subjects' in df.columns and df['failed_subjects'].astype(str).str.strip().ne('').any():
                    failed_rows = df[df['failed_subjects'].astype(str).str.contains(subject_name, regex=False, na=False)]
                else:
                    failed_rows = df[df[col] < passing]
                failed = len(failed_rows)
                total_atkt = failed_rows['atkt_count'].sum()
                fail_rate = (failed / total * 100) if total > 0 else 0
                
                self.atkt_table.setItem(i, 0, QTableWidgetItem(subject_name))
                self.atkt_table.setItem(i, 1, QTableWidgetItem(str(len(df))))
                self.atkt_table.setItem(i, 2, QTableWidgetItem(str(failed)))
                self.atkt_table.setItem(i, 3, QTableWidgetItem(str(total_atkt)))
                
                rate_item = QTableWidgetItem(f"{fail_rate:.1f}%")
                if fail_rate > 30:
                    rate_item.setForeground(QColor("#dc2626"))
                elif fail_rate > 10:
                    rate_item.setForeground(QColor("#f97316"))
                else:
                    rate_item.setForeground(QColor("#16a34a"))
                self.atkt_table.setItem(i, 4, rate_item)
        
        self.atkt_table.resizeColumnsToContents()

    # ── Subject Topper ──────────────────────────────────────────────────────
    def refresh_subject_topper(self, show_all=True):
        df = self.db.get_all_data()
        if df.empty:
            QMessageBox.information(self, "No Data", "Please upload a PDF first.")
            return

        self.topper_table.setRowCount(0)
        rows = []

        if show_all:
            # Show topper for EVERY subject
            for i in range(7):
                col = f'sub_{i+1}'
                if col not in df.columns: continue
                sub_name = self.get_subject_name(i)
                best_idx = df[col].idxmax()
                row = df.loc[best_idx]
                rows.append((1, str(row['roll_no']), str(row['name']),
                              sub_name, int(row[col]),
                              str(row['sgpa']), str(row['status'])))
        else:
            # Show top 5 for selected subject
            idx = self.combo_topper.currentIndex()
            col = f'sub_{idx+1}'
            if col not in df.columns: return
            sub_name = self.get_subject_name(idx)
            top5 = df.nlargest(5, col)
            for rank, (_, row) in enumerate(top5.iterrows(), start=1):
                rows.append((rank, str(row['roll_no']), str(row['name']),
                              sub_name, int(row[col]),
                              str(row['sgpa']), str(row['status'])))

        self.topper_table.setRowCount(len(rows))
        rank_colors = ["#f97316", "#64748b", "#ca8a04"]
        for r, (rank, roll, name, subj, marks, sgpa, status) in enumerate(rows):
            items = [str(rank), roll, name, subj, str(marks), sgpa, status]
            for c, txt in enumerate(items):
                item = QTableWidgetItem(txt)
                if c == 0 and rank <= 3:
                    item.setForeground(QColor(rank_colors[rank - 1]))
                    item.setFont(QFont("Segoe UI", 11, QFont.Bold))
                if c == 6:
                    item.setForeground(QColor("#16a34a") if status == "PASS" else QColor("#dc2626"))
                self.topper_table.setItem(r, c, item)

    # ── SGPA Analysis Page ───────────────────────────────────────────────────
    def create_sgpa_analysis_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        pg_title = QLabel("📉  SGPA & Grade Analysis")
        pg_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0c1a3a;")
        layout.addWidget(pg_title)

        # KPI cards
        kpi_row = QHBoxLayout(); kpi_row.setSpacing(14)
        self.kpi_avg_sgpa  = KPICard("Average SGPA",   "—", "#7c3aed", "📉")
        self.kpi_max_sgpa  = KPICard("Highest SGPA",   "—", "#0d9488", "🏆")
        self.kpi_min_sgpa  = KPICard("Lowest SGPA",    "—", "#dc2626", "📌")
        self.kpi_o_grade   = KPICard("O Grade (≥ 9.0)","—", "#f97316", "⭐")
        for card in [self.kpi_avg_sgpa, self.kpi_max_sgpa, self.kpi_min_sgpa, self.kpi_o_grade]:
            kpi_row.addWidget(card)
        layout.addLayout(kpi_row)

        # SGPA distribution chart
        chart_outer, _, _, chart_inner = make_section_card("📊  SGPA Distribution Chart", collapsible=False)
        self.sgpa_canvas = MplCanvas(self, width=10, height=4.0)
        chart_inner.addWidget(self.sgpa_canvas)
        layout.addWidget(chart_outer)

        # Grade breakdown table
        grade_outer, _, _, grade_inner = make_section_card("🎓  Grade-wise Breakdown", collapsible=False)
        self.grade_table = QTableWidget()
        self.grade_table.setColumnCount(3)
        self.grade_table.setHorizontalHeaderLabels(["Grade", "No. of Students", "Percentage (%)"])
        self.setup_table(self.grade_table, row_height=34)
        self.grade_table.setMaximumHeight(280)
        gh = self.grade_table.horizontalHeader()
        gh.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        gh.setSectionResizeMode(1, QHeaderView.Stretch)
        gh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        grade_inner.addWidget(self.grade_table)
        layout.addWidget(grade_outer)

        layout.addStretch()
        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def refresh_sgpa_analysis(self):
        df = self.db.get_all_data()
        if df.empty: return
        try:
            sgpa_vals = pd.to_numeric(df['sgpa'], errors='coerce').dropna()
            sgpa_vals = sgpa_vals[sgpa_vals > 0]
            if sgpa_vals.empty: return

            avg_s = round(sgpa_vals.mean(), 2)
            max_s = round(sgpa_vals.max(), 2)
            min_s = round(sgpa_vals.min(), 2)
            o_cnt = int((sgpa_vals >= 9.0).sum())

            self.kpi_avg_sgpa.lbl_value.setText(str(avg_s))
            self.kpi_max_sgpa.lbl_value.setText(str(max_s))
            self.kpi_min_sgpa.lbl_value.setText(str(min_s))
            self.kpi_o_grade.lbl_value.setText(str(o_cnt))

            # SGPA histogram
            self.sgpa_canvas.fig.patch.set_facecolor('#ffffff')
            self.sgpa_canvas.axes.clear()
            self.sgpa_canvas.axes.set_facecolor('#f8fafc')
            self.sgpa_canvas.axes.hist(
                sgpa_vals, bins=20, color='#7c3aed',
                edgecolor='#ffffff', linewidth=1.2, alpha=0.85
            )
            self.sgpa_canvas.axes.axvline(avg_s, color='#f97316', linestyle='--',
                                           linewidth=2, label=f'Avg SGPA: {avg_s}')
            self.sgpa_canvas.axes.set_title("SGPA Distribution", fontsize=12,
                                             fontweight='bold', color='#0c1a3a')
            self.sgpa_canvas.axes.set_xlabel("SGPA", color='#64748b')
            self.sgpa_canvas.axes.set_ylabel("No. of Students", color='#64748b')
            self.sgpa_canvas.axes.grid(axis='y', color='#e2e8f0', linewidth=0.8)
            self.sgpa_canvas.axes.legend(frameon=False)
            self.sgpa_canvas.axes.spines['top'].set_visible(False)
            self.sgpa_canvas.axes.spines['right'].set_visible(False)
            self.sgpa_canvas.axes.tick_params(colors='#64748b')
            self.sgpa_canvas.draw()

            # Grade table
            gc = df['grade'].value_counts().sort_index()
            total = len(df)
            self.grade_table.setRowCount(len(gc))
            for i, (grade, count) in enumerate(gc.items()):
                pct = round(count / total * 100, 1) if total else 0
                self.grade_table.setItem(i, 0, QTableWidgetItem(str(grade)))
                self.grade_table.setItem(i, 1, QTableWidgetItem(str(count)))
                pct_item = QTableWidgetItem(f"{pct}%")
                if grade == 'O':
                    pct_item.setForeground(QColor("#0d9488"))
                elif grade == 'F':
                    pct_item.setForeground(QColor("#dc2626"))
                self.grade_table.setItem(i, 2, pct_item)
        except Exception as e:
            print(f"SGPA refresh err: {e}")

    # ── Student Lookup Page ──────────────────────────────────────────────────
    def create_student_lookup_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        pg_title = QLabel("🔍  Student Lookup")
        pg_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0c1a3a;")
        layout.addWidget(pg_title)

        # Search bar
        search_outer, _, _, search_inner = make_section_card("🔎  Search Student", collapsible=False)
        search_row = QHBoxLayout()
        self.lookup_search = QLineEdit()
        self.lookup_search.setPlaceholderText("Enter Roll No or Student Name and press Enter / click Search...")
        self.lookup_search.setFixedHeight(40)
        self.lookup_search.returnPressed.connect(self.do_student_lookup)
        search_row.addWidget(self.lookup_search)
        lookup_btn = QPushButton("🔍  Search")
        lookup_btn.setFixedWidth(130)
        lookup_btn.setCursor(QCursor(Qt.PointingHandCursor))
        lookup_btn.clicked.connect(self.do_student_lookup)
        search_row.addWidget(lookup_btn)
        search_inner.addLayout(search_row)
        layout.addWidget(search_outer)

        # Results table
        result_outer, _, _, result_inner = make_section_card("📋  Search Results", collapsible=False)
        self.lookup_status_lbl = QLabel("Enter a roll number or name above and press Search.")
        self.lookup_status_lbl.setStyleSheet("color: #64748b; font-size: 13px; padding: 10px 0px;")
        result_inner.addWidget(self.lookup_status_lbl)

        self.lookup_table = QTableWidget()
        self.lookup_table.setColumnCount(7)
        self.lookup_table.setHorizontalHeaderLabels([
            "Roll No", "Name", "Total", "SGPA", "Grade", "Status", "ATKT"
        ])
        self.setup_table(self.lookup_table, row_height=36)
        self.lookup_table.setMinimumHeight(200)
        lh = self.lookup_table.horizontalHeader()
        lh.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(1, QHeaderView.Stretch)
        lh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        result_inner.addWidget(self.lookup_table)
        layout.addWidget(result_outer)

        # Detail card — shows full marks of selected student
        detail_outer, _, _, detail_inner = make_section_card("📄  Detailed Result Card", collapsible=False)
        self.lookup_detail_lbl = QLabel("Select a student from the results above to see full marks.")
        self.lookup_detail_lbl.setStyleSheet("color: #64748b; font-size: 13px; padding: 10px 0px;")
        self.lookup_detail_lbl.setWordWrap(True)
        self.lookup_detail_lbl.setAlignment(Qt.AlignCenter)
        detail_inner.addWidget(self.lookup_detail_lbl)

        self.lookup_marks_table = QTableWidget()
        self.lookup_marks_table.setColumnCount(3)
        self.lookup_marks_table.setHorizontalHeaderLabels(["Subject", "Marks", "Status"])
        self.setup_table(self.lookup_marks_table, row_height=34)
        self.lookup_marks_table.setMinimumHeight(200)       # scrollable
        self.lookup_marks_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        mh = self.lookup_marks_table.horizontalHeader()
        mh.setSectionResizeMode(0, QHeaderView.Stretch)
        mh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        mh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.lookup_marks_table.hide()
        detail_inner.addWidget(self.lookup_marks_table)
        layout.addWidget(detail_outer)

        # Connect table selection to detail view
        self.lookup_table.itemSelectionChanged.connect(self.on_lookup_selection)

        layout.addStretch()
        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def do_student_lookup(self):
        query = self.lookup_search.text().strip()
        if not query:
            QMessageBox.information(self, "Search", "Please enter a Roll No or Name.")
            return
        df = self.db.get_all_data()
        if df.empty:
            self.lookup_status_lbl.setText("No data loaded. Please upload a PDF first.")
            return
        q = query.lower()
        mask = (df['name'].astype(str).str.lower().str.contains(q, na=False) |
                df['roll_no'].astype(str).str.lower().str.contains(q, na=False))
        results = df[mask]
        self.lookup_table.setRowCount(0)
        self.lookup_marks_table.hide()
        self.lookup_detail_lbl.setText("Select a student from results above to see full marks.")
        self.lookup_detail_lbl.show()
        if results.empty:
            self.lookup_status_lbl.setText(f"No results found for: '{query}'")
            return
        self.lookup_status_lbl.setText(f"Found {len(results)} student(s) matching '{query}':")
        self.lookup_table.setRowCount(len(results))
        self._lookup_df = results.reset_index(drop=True)
        for r, (_, row) in enumerate(results.iterrows()):
            vals = [str(row['roll_no']), str(row['name']), str(row['total']),
                    str(row['sgpa']), str(row.get('grade','—')),
                    str(row['status']), str(row['atkt_count'])]
            for c, v in enumerate(vals):
                item = QTableWidgetItem(v)
                if c == 5:
                    item.setForeground(QColor("#16a34a") if v == "PASS" else QColor("#dc2626"))
                self.lookup_table.setItem(r, c, item)

    def on_lookup_selection(self):
        rows = self.lookup_table.selectedItems()
        if not rows or not hasattr(self, '_lookup_df'): return
        r = self.lookup_table.currentRow()
        if r < 0 or r >= len(self._lookup_df): return
        row = self._lookup_df.iloc[r]
        self.lookup_detail_lbl.hide()
        self.lookup_marks_table.show()
        self.lookup_marks_table.setRowCount(7)
        for i in range(7):
            col = f'sub_{i+1}'
            sub_name = self.get_subject_name(i)
            marks = int(row.get(col, 0)) if col in row else 0
            try:
                passing = final_result.EXT_PASSING_MIN[i] if PARSER_LOADED else 18
            except Exception:
                passing = 18
            status = "PASS" if marks >= passing else "FAIL"
            self.lookup_marks_table.setItem(i, 0, QTableWidgetItem(sub_name))
            m_item = QTableWidgetItem(str(marks))
            m_item.setTextAlignment(Qt.AlignCenter)
            self.lookup_marks_table.setItem(i, 1, m_item)
            s_item = QTableWidgetItem(status)
            s_item.setForeground(QColor("#16a34a") if status == "PASS" else QColor("#dc2626"))
            self.lookup_marks_table.setItem(i, 2, s_item)

    # ── Settings & Print ─────────────────────────────────────────────────────
    def show_report_dialog(self):
        """Settings page pe Report dialog kholo — type aur format choose karo."""
        if not self.data_loaded:
            QMessageBox.warning(self, "No Data", "Please upload a PDF first.")
            return
        stats = self.db.get_stats()
        # Get current Top-N value from the Top N Rankers page spinner
        top_n = self.spin_n.value() if hasattr(self, 'spin_n') else 10
        dlg = ReportTypeDialog(self, stats, top_n=top_n)
        if dlg.exec() != QDialog.Accepted:
            return
        rtype, fmt, top_n = dlg.get_selection()
        df = self.db.get_all_data()

        if fmt == 'excel':
            if rtype == 'pass':
                self.export_to_excel(df[df['status'] == 'PASS'])
            elif rtype == 'fail':
                self.export_to_excel(df[df['status'] == 'FAIL'])
            elif rtype == 'top10':
                try:
                    sorted_all = self.sort_by_sgpa_then_roll_no(df).reset_index(drop=True)
                except Exception:
                    sorted_all = df.reset_index(drop=True)
                all_ranks = self.sgpa_rank_numbers(sorted_all)
                top_df = sorted_all[[r <= top_n for r in all_ranks]].reset_index(drop=True)
                self.export_to_excel(top_df, f"Top{top_n}_Rankers.xlsx")
            elif rtype == 'subject':
                self.export_to_excel(df, "Subject_Wise_Report.xlsx")
            else:
                self.export_to_excel(df)
        else:
            # PDF — comprehensive for 'all', dedicated for others
            _default_pdf = os.path.join(BASE_DIR, "Student_Report.pdf")
            path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", _default_pdf, "PDF Files (*.pdf)")
            if not path:
                return
            if not path.lower().endswith(".pdf"):
                path += ".pdf"
            # Ensure parent directory exists
            os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
            try:
                if rtype == 'all':
                    self.generate_comprehensive_pdf(df, path)
                elif rtype == 'pass':
                    self.generate_comprehensive_pdf(df[df['status'] == 'PASS'].reset_index(drop=True), path, section='pass')
                elif rtype == 'fail':
                    self.generate_comprehensive_pdf(df[df['status'] == 'FAIL'].reset_index(drop=True), path, section='fail')
                elif rtype == 'top10':
                    self.generate_comprehensive_pdf(df, path, section='top10', top_n=top_n)
                elif rtype == 'subject':
                    self.generate_comprehensive_pdf(df, path, section='subject')
                QMessageBox.information(self, "✅ Success", f"PDF Report saved successfully:\n{path}")
            except PermissionError:
                # Auto-retry with a timestamped filename so user doesn't have to close PDF viewer
                import time as _time
                ts = _time.strftime("%H%M%S")
                base, ext = os.path.splitext(path)
                new_path = f"{base}_{ts}{ext}"
                retry = QMessageBox.question(
                    self, "File Locked 🔒",
                    f"Cannot save — the file is already open in a PDF viewer:\n{path}\n\n"
                    f"Click Yes to save as a new file:\n{new_path}",
                    QMessageBox.Yes | QMessageBox.No
                )
                if retry == QMessageBox.Yes:
                    try:
                        if rtype == 'all':
                            self.generate_comprehensive_pdf(df, new_path)
                        elif rtype == 'pass':
                            self.generate_comprehensive_pdf(df[df['status'] == 'PASS'].reset_index(drop=True), new_path, section='pass')
                        elif rtype == 'fail':
                            self.generate_comprehensive_pdf(df[df['status'] == 'FAIL'].reset_index(drop=True), new_path, section='fail')
                        elif rtype == 'top10':
                            self.generate_comprehensive_pdf(df, new_path, section='top10', top_n=top_n)
                        elif rtype == 'subject':
                            self.generate_comprehensive_pdf(df, new_path, section='subject')
                        QMessageBox.information(self, "✅ Saved", f"PDF saved successfully:\n{new_path}")
                    except Exception as e2:
                        QMessageBox.critical(self, "PDF Error", f"Could not generate PDF:\n{e2}")
            except Exception as e:
                import traceback; traceback.print_exc()
                QMessageBox.critical(self, "PDF Error", f"Could not generate PDF:\n{e}")

    # ─────────────────────────────────────────────────────────────────────────
    # COMPREHENSIVE PDF GENERATOR
    # ─────────────────────────────────────────────────────────────────────────
    def _get_first_page_text(self, pdf_path):
        if not pdf_path or final_result is None:
            return ""

        pdfplumber_lib = getattr(final_result, 'pdfplumber', None)
        if pdfplumber_lib is None:
            return ""

        try:
            with pdfplumber_lib.open(pdf_path) as pdf:
                if not pdf.pages:
                    return ""
                return pdf.pages[0].extract_text() or ""
        except Exception:
            return ""

    def _select_pdf_header_path(self):
        first_page_text = self._get_first_page_text(self.current_pdf_path)
        if "(0349)" in first_page_text:
            return SASCMA_PDF_HEADER_PATH
        return UNIVERSITY_PDF_HEADER_PATH

    def generate_comprehensive_pdf(self, df, path, section='all', top_n=10):
        """
        Multi-section PDF with:
          - Header image on every page
          - Section banner (dark navy bg + white text) below header
          - Footer: course name (left) + Page X (right)
        Sections (when section='all'):
          1. All Students complete result table
          2. Top N Rankers
          3. Failed / ATKT Students
          4. Subject-wise Performance Summary
        """
        from reportlab.platypus import (Paragraph, Table, TableStyle, Spacer,
                                         Image as RLImage, PageBreak, KeepTogether)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.platypus.doctemplate import BaseDocTemplate, PageTemplate
        from reportlab.platypus.frames import Frame

        page_w, page_h = A4          # Portrait A4 (default)
        margin = 0.5 * inch
        pdf_header_path = self._select_pdf_header_path()
        header_img_exists = os.path.exists(pdf_header_path)
        # Auto-calculate header height from actual image aspect ratio
        if header_img_exists:
            try:
                from reportlab.lib.utils import ImageReader
                _ir = ImageReader(pdf_header_path)
                _iw, _ih = _ir.getSize()
                # Draw full-width: height = (img_h / img_w) * page_w, cap at 2.0" to stay safe
                HDR_H = min((_ih / _iw) * page_w, 2.0 * inch)
            except Exception:
                HDR_H = 1.2 * inch
        else:
            HDR_H = 0.0
        FTR_H = 0.35 * inch

        # ── Page callback ──────────────────────────────────────────────────
        def draw_page(canvas, doc):
            canvas.saveState()
            # Header image — drawn full-width, exact aspect ratio height
            if header_img_exists:
                canvas.drawImage(
                    pdf_header_path,
                    0, page_h - HDR_H,          # x=0 (flush left), y = top of content
                    width=page_w,               # stretch to full page width
                    height=HDR_H,               # exact calculated height
                    mask='auto'
                )
                # Thin separator line below header
                canvas.setStrokeColor(colors.HexColor('#0c1a3a'))
                canvas.setLineWidth(1.0)
                canvas.line(0, page_h - HDR_H, page_w, page_h - HDR_H)
            # Footer separator
            canvas.setStrokeColor(colors.HexColor('#cbd5e1'))
            canvas.setLineWidth(0.4)
            canvas.line(margin, margin + 0.20*inch, page_w - margin, margin + 0.20*inch)
            # Footer text
            canvas.setFont('Helvetica', 7.5)
            canvas.setFillColor(colors.HexColor('#64748b'))
            canvas.drawString(margin, margin + 0.06*inch, COURSE_NAME)
            canvas.drawRightString(page_w - margin, margin + 0.06*inch, f"Page {doc.page}")
            canvas.restoreState()

        # ── Frame & template ───────────────────────────────────────────────
        frame_x = margin
        frame_y = margin + FTR_H
        frame_w = page_w - 2 * margin
        frame_h = page_h - HDR_H - margin - FTR_H - 0.06*inch

        frame = Frame(frame_x, frame_y, frame_w, frame_h, id='main',
                      leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
        template = PageTemplate(id='std', frames=[frame], onPage=draw_page)

        doc = BaseDocTemplate(
            path, pagesize=A4,       # Portrait A4
            leftMargin=margin, rightMargin=margin,
            topMargin=HDR_H + 0.06*inch, bottomMargin=margin + FTR_H,
        )
        doc.addPageTemplates([template])

        # ── Styles ─────────────────────────────────────────────────────────
        styles = getSampleStyleSheet()
        CW = frame_w  # content width

        def banner(title):
            """Section title banner — dark navy bg, white bold text (like image 3)."""
            ps = ParagraphStyle('bn', fontName='Helvetica-Bold', fontSize=13,
                                textColor=colors.white, leftIndent=6)
            t = Table([[Paragraph(title, ps)]], colWidths=[CW])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#0c1a3a')),
                ('TOPPADDING',    (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
                ('LEFTPADDING',   (0, 0), (-1, -1), 14),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 14),
            ]))
            return t

        def hdr_style():
            return ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=7,
                                  textColor=colors.white)

        def body_style():
            return ParagraphStyle('b', fontName='Helvetica', fontSize=7,
                                  textColor=colors.HexColor('#1e293b'))

        TBL_BASE = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c1a3a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f0fdfa')]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]

        # ── Helper: build subject names ────────────────────────────────────
        sub_names = [self.get_subject_name(i) for i in range(7)]

        story = []

        # ── For pass/fail: show a simple filtered list, then return ────────
        if section in ('pass', 'fail'):
            sec_title = "✅  Passed Students" if section == 'pass' else "❌  Failed / ATKT Students"
            story.append(banner(sec_title))
            story.append(Spacer(1, 0.1*inch))
            hdrs = ['Roll No', 'Student Name'] + sub_names + ['Total', 'SGPA', 'Grade', 'Status', 'ATKT']
            fixed = 0.6 + 1.7 + 0.5 + 0.5 + 0.5 + 0.65 + 0.5
            sub_w = (CW - fixed * inch) / 7
            col_ws = ([0.6*inch, 1.7*inch] + [sub_w]*7 +
                      [0.5*inch, 0.5*inch, 0.5*inch, 0.65*inch, 0.5*inch])
            data = [[Paragraph(h, hdr_style()) for h in hdrs]]
            ts_extra = list(TBL_BASE)
            for r_idx, (_, row) in enumerate(df.iterrows(), start=1):
                cells = [str(row.get('roll_no', '')),
                         Paragraph(str(row.get('name', '')), body_style())]
                for i in range(7):
                    cells.append(str(row.get(f'sub_{i+1}', '')))
                cells += [str(row.get('total', '')), str(row.get('sgpa', '')),
                          str(row.get('grade', '')), str(row.get('status', '')),
                          str(row.get('atkt_count', 0))]
                data.append(cells)
                st_col = len(hdrs) - 2
                col_c = (colors.HexColor('#16a34a') if str(row.get('status', '')) == 'PASS'
                         else colors.HexColor('#dc2626'))
                ts_extra.append(('TEXTCOLOR', (st_col, r_idx), (st_col, r_idx), col_c))
                if str(row.get('status', '')) == 'FAIL':
                    ts_extra.append(('FONTNAME', (st_col, r_idx), (st_col, r_idx), 'Helvetica-Bold'))
            t_filt = Table(data, colWidths=col_ws, repeatRows=1)
            t_filt.setStyle(TableStyle(ts_extra))
            story.append(t_filt)
            doc.build(story)
            return

        # ── Top 10 Rankers section ────────────────────────────────────────
        if section == 'top10':
            story.append(banner(f"🏆  Top {top_n} Rankers — Sorted by SGPA"))
            story.append(Spacer(1, 0.12 * inch))

            try:
                sorted_all = self.sort_by_sgpa_then_roll_no(df).reset_index(drop=True)
            except Exception:
                sorted_all = df.reset_index(drop=True)
            all_ranks_pdf = self.sgpa_rank_numbers(sorted_all)
            top_df = sorted_all[[r <= top_n for r in all_ranks_pdf]].reset_index(drop=True)

            total_st = len(df)
            pass_st  = int((df['status'] == 'PASS').sum()) if 'status' in df.columns else 0
            fail_st  = total_st - pass_st
            try:
                all_sgpa = pd.to_numeric(df['sgpa'], errors='coerce').dropna()
                all_sgpa = all_sgpa[all_sgpa > 0]
                class_avg   = f"{all_sgpa.mean():.2f}" if len(all_sgpa) else "—"
                topper_sgpa = str(top_df.iloc[0].get('sgpa', '—')) if len(top_df) else "—"
            except Exception:
                class_avg, topper_sgpa = "—", "—"

            kpi_ps_t  = ParagraphStyle('kpit',  fontName='Helvetica',      fontSize=7.5, textColor=colors.HexColor('#64748b'))
            kpi_val_t = ParagraphStyle('kpivt', fontName='Helvetica-Bold', fontSize=13,  textColor=colors.HexColor('#0c1a3a'))
            kpi_g_t   = ParagraphStyle('kpigt', fontName='Helvetica-Bold', fontSize=13,  textColor=colors.HexColor('#16a34a'))
            kpi_r_t   = ParagraphStyle('kpirt', fontName='Helvetica-Bold', fontSize=13,  textColor=colors.HexColor('#dc2626'))
            kpi_p_t   = ParagraphStyle('kpipt', fontName='Helvetica-Bold', fontSize=13,  textColor=colors.HexColor('#7c3aed'))
            kpi_o_t   = ParagraphStyle('kpiot', fontName='Helvetica-Bold', fontSize=13,  textColor=colors.HexColor('#d97706'))

            def _kc(label, val, ps):
                return [Paragraph(val, ps), Paragraph(label, kpi_ps_t)]

            kpi_w = CW / 5
            kd = [[_kc("Total Students", str(total_st), kpi_val_t),
                   _kc("Passed",         str(pass_st),  kpi_g_t),
                   _kc("Failed",         str(fail_st),  kpi_r_t),
                   _kc("Class Avg SGPA", class_avg,     kpi_p_t),
                   _kc("Topper SGPA",    topper_sgpa,   kpi_o_t)]]
            kt = Table(kd, colWidths=[kpi_w] * 5)
            kt.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, -1), colors.HexColor('#fefce8')),
                ('BOX',           (0, 0), (-1, -1), 0.8, colors.HexColor('#d97706')),
                ('INNERGRID',     (0, 0), (-1, -1), 0.4, colors.HexColor('#fde68a')),
                ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING',    (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
            ]))
            story.append(kt)
            story.append(Spacer(1, 0.15 * inch))

            medals = {1: "🥇 1st", 2: "🥈 2nd", 3: "🥉 3rd"}
            t10_hdrs = ['Rank', 'Roll No', 'Student Name', 'College / Institution', 'Total', 'SGPA', 'Grade', 'Status']
            # Fixed widths: rank+roll+total+sgpa+grade+status = 0.55+0.6+0.65+0.6+0.55+0.65 = 3.60"
            fixed_w  = (0.55 + 0.6 + 0.65 + 0.6 + 0.55 + 0.65) * inch
            remaining = CW - fixed_w
            name_w_t10   = remaining * 0.40   # 40% of remaining for name
            college_w_t10 = remaining * 0.60  # 60% for college
            t10_cws = [0.55*inch, 0.6*inch, name_w_t10, college_w_t10, 0.65*inch, 0.6*inch, 0.55*inch, 0.65*inch]

            t10_data = [[Paragraph(h, hdr_style()) for h in t10_hdrs]]
            ts_t10   = list(TBL_BASE)

            rank_numbers = all_ranks_pdf[:len(top_df)]
            for row_no, ((_, row), rank_no) in enumerate(zip(top_df.iterrows(), rank_numbers), start=1):
                rank_text = medals.get(rank_no, str(rank_no))
                st = str(row.get('status', ''))
                college_val = str(row.get('college', '')).strip() or '—'
                cells = [rank_text, str(row.get('roll_no', '')),
                         Paragraph(str(row.get('name', '')), body_style()),
                         Paragraph(college_val, body_style()),
                         str(row.get('total', '')), str(row.get('sgpa', '')),
                         str(row.get('grade', '')), st]
                t10_data.append(cells)
                if rank_no == 1:
                    ts_t10.append(('BACKGROUND', (0, row_no), (-1, row_no), colors.HexColor('#fef9c3')))
                    ts_t10.append(('FONTNAME',   (0, row_no), (-1, row_no), 'Helvetica-Bold'))
                elif rank_no == 2:
                    ts_t10.append(('BACKGROUND', (0, row_no), (-1, row_no), colors.HexColor('#f1f5f9')))
                    ts_t10.append(('FONTNAME',   (0, row_no), (-1, row_no), 'Helvetica-Bold'))
                elif rank_no == 3:
                    ts_t10.append(('BACKGROUND', (0, row_no), (-1, row_no), colors.HexColor('#fff7ed')))
                    ts_t10.append(('FONTNAME',   (0, row_no), (-1, row_no), 'Helvetica-Bold'))
                col_c = colors.HexColor('#16a34a') if st == 'PASS' else colors.HexColor('#dc2626')
                ts_t10.append(('TEXTCOLOR', (t10_hdrs.index('Status'), row_no),
                                             (t10_hdrs.index('Status'), row_no), col_c))

            tbl_t10 = Table(t10_data, colWidths=t10_cws, repeatRows=1)
            tbl_t10.setStyle(TableStyle(ts_t10))
            story.append(tbl_t10)
            doc.build(story)
            return


        # ── Subject-wise Performance section ─────────────────────────────
        if section == 'subject':
            story.append(banner("📊  Subject-wise Performance Report"))
            story.append(Spacer(1, 0.12 * inch))

            s_hdrs2 = ['Subject', 'Total', 'Pass', 'Fail', 'Avg', 'Highest', 'Lowest', 'Pass %', 'Topper', 'Marks']
            s_fixed2 = 0.5 + 0.5 + 0.5 + 0.5 + 0.55 + 0.55 + 0.55 + 1.3 + 0.5
            s_name_w2 = CW - s_fixed2 * inch
            s_cws2 = [s_name_w2, 0.5*inch, 0.5*inch, 0.5*inch,
                      0.5*inch, 0.55*inch, 0.55*inch, 0.55*inch, 1.3*inch, 0.5*inch]

            sub_data2 = [[Paragraph(h, hdr_style()) for h in s_hdrs2]]
            ts_sub2   = list(TBL_BASE)

            for i in range(7):
                col = f'sub_{i+1}'
                if col not in df.columns:
                    continue
                sname  = sub_names[i]
                marks  = pd.to_numeric(df[col], errors='coerce')
                valid  = marks[marks > 0]
                tot_s  = len(df)
                try:
                    passing = final_result.EXT_PASSING_MIN[i] if PARSER_LOADED else 18
                except Exception:
                    passing = 18
                if 'failed_subjects' in df.columns and df['failed_subjects'].astype(str).str.strip().ne('').any():
                    fail_cnt2 = int(df['failed_subjects'].astype(str).str.contains(sname, regex=False, na=False).sum())
                else:
                    fail_cnt2 = int((marks < passing).sum())
                pass_cnt2 = tot_s - fail_cnt2
                avg2      = f"{valid.mean():.1f}" if len(valid) else "—"
                high2     = str(int(valid.max()))  if len(valid) else "—"
                low2      = str(int(valid.min()))  if len(valid) else "—"
                pct2      = f"{pass_cnt2/tot_s*100:.1f}%" if tot_s else "—"
                if len(valid):
                    ti2 = marks.idxmax()
                    top_name2 = str(df.loc[ti2, 'name']) if ti2 in df.index else "—"
                    top_mark2 = str(int(marks.loc[ti2]))
                else:
                    top_name2, top_mark2 = "—", "—"

                ri2 = len(sub_data2)
                sub_data2.append([
                    Paragraph(sname, body_style()),
                    str(tot_s), str(pass_cnt2), str(fail_cnt2),
                    avg2, high2, low2, pct2,
                    Paragraph(top_name2, body_style()), top_mark2
                ])
                if fail_cnt2 > 0:
                    ts_sub2.append(('TEXTCOLOR', (3, ri2), (3, ri2), colors.HexColor('#dc2626')))
                    ts_sub2.append(('FONTNAME',  (3, ri2), (3, ri2), 'Helvetica-Bold'))
                ts_sub2.append(('TEXTCOLOR', (2, ri2), (2, ri2), colors.HexColor('#16a34a')))

            tbl_sub2 = Table(sub_data2, colWidths=s_cws2, repeatRows=1)
            tbl_sub2.setStyle(TableStyle(ts_sub2))
            story.append(tbl_sub2)
            story.append(Spacer(1, 0.20 * inch))

            # Subject toppers block
            story.append(banner("🏅  Subject Toppers"))
            story.append(Spacer(1, 0.10 * inch))

            top2_hdrs = ['Subject', 'Topper Name', 'Roll No', 'Marks', 'SGPA', 'Status']
            top2_fixed = 0.6 + 1.5 + 0.55 + 0.55 + 0.65
            top2_name_w = CW - top2_fixed * inch
            top2_cws = [top2_name_w, 1.5*inch, 0.6*inch, 0.55*inch, 0.55*inch, 0.65*inch]

            top2_data = [[Paragraph(h, hdr_style()) for h in top2_hdrs]]
            ts_top2   = list(TBL_BASE)

            for i in range(7):
                col = f'sub_{i+1}'
                if col not in df.columns:
                    continue
                sname  = sub_names[i]
                marks  = pd.to_numeric(df[col], errors='coerce')
                valid  = marks[marks > 0]
                if not len(valid):
                    continue
                ti3 = marks.idxmax()
                if ti3 not in df.index:
                    continue
                t_row3  = df.loc[ti3]
                t_name3 = str(t_row3.get('name', '—'))
                t_roll3 = str(t_row3.get('roll_no', '—'))
                t_mark3 = str(int(marks.loc[ti3]))
                t_sgpa3 = str(t_row3.get('sgpa', '—'))
                t_st3   = str(t_row3.get('status', '—'))
                ri3     = len(top2_data)
                top2_data.append([
                    Paragraph(sname, body_style()),
                    Paragraph(t_name3, body_style()),
                    t_roll3, t_mark3, t_sgpa3, t_st3
                ])
                col_c3 = colors.HexColor('#16a34a') if t_st3 == 'PASS' else colors.HexColor('#dc2626')
                ts_top2.append(('TEXTCOLOR', (5, ri3), (5, ri3), col_c3))
                if ri3 % 2 == 1:
                    ts_top2.append(('BACKGROUND', (0, ri3), (-1, ri3), colors.HexColor('#fefce8')))

            tbl_top2 = Table(top2_data, colWidths=top2_cws, repeatRows=1)
            tbl_top2.setStyle(TableStyle(ts_top2))
            story.append(tbl_top2)
            doc.build(story)
            return

        # ═══════════════════════════════════════════════════════════════════
        # 'all' mode — Section 1: Top N Rankers  (All Students removed)
        # ═══════════════════════════════════════════════════════════════════
        story.append(banner("🏆  Top Rankers (Sorted by SGPA)"))
        story.append(Spacer(1, 0.1*inch))


        top_df = df.copy()
        try:
            top_df = self.sort_by_sgpa_then_roll_no(top_df).reset_index(drop=True)
        except Exception:
            pass

        rank_hdrs = ['Rank', 'Roll No', 'Student Name', 'Total', 'SGPA', 'Grade', 'Status', 'ATKT']
        rank_cws = [0.38*inch, 0.6*inch, 2.0*inch, 0.55*inch,
                    0.55*inch, 0.5*inch, 0.6*inch, 0.46*inch]

        rank_data = [[Paragraph(h, hdr_style()) for h in rank_hdrs]]
        ts_rank = list(TBL_BASE)

        rank_numbers = self.sgpa_rank_numbers(top_df)
        for row_no, ((_, row), rank_no) in enumerate(zip(top_df.iterrows(), rank_numbers), start=1):
            cells = [str(rank_no), str(row.get('roll_no', '')),
                     Paragraph(str(row.get('name', '')), body_style()),
                     str(row.get('total', '')), str(row.get('sgpa', '')),
                     str(row.get('grade', '')), str(row.get('status', '')),
                     str(row.get('atkt_count', 0))]
            rank_data.append(cells)
            # Gold/silver/bronze for top 3
            if rank_no == 1:
                ts_rank.append(('BACKGROUND', (0, row_no), (-1, row_no), colors.HexColor('#fef9c3')))
                ts_rank.append(('TEXTCOLOR', (0, row_no), (0, row_no), colors.HexColor('#ca8a04')))
                ts_rank.append(('FONTNAME', (0, row_no), (-1, row_no), 'Helvetica-Bold'))
            elif rank_no == 2:
                ts_rank.append(('BACKGROUND', (0, row_no), (-1, row_no), colors.HexColor('#f1f5f9')))
                ts_rank.append(('TEXTCOLOR', (0, row_no), (0, row_no), colors.HexColor('#475569')))
                ts_rank.append(('FONTNAME', (0, row_no), (-1, row_no), 'Helvetica-Bold'))
            elif rank_no == 3:
                ts_rank.append(('BACKGROUND', (0, row_no), (-1, row_no), colors.HexColor('#fff7ed')))
                ts_rank.append(('TEXTCOLOR', (0, row_no), (0, row_no), colors.HexColor('#c2410c')))
                ts_rank.append(('FONTNAME', (0, row_no), (-1, row_no), 'Helvetica-Bold'))
            st_col = rank_hdrs.index('Status')
            col_c = colors.HexColor('#16a34a') if str(row.get('status','')) == 'PASS' else colors.HexColor('#dc2626')
            ts_rank.append(('TEXTCOLOR', (st_col, row_no), (st_col, row_no), col_c))

        t2 = Table(rank_data, colWidths=rank_cws, repeatRows=1)
        t2.setStyle(TableStyle(ts_rank))
        story.append(t2)

        # ═══════════════════════════════════════════════════════════════════
        # SECTION 3: Failed Students
        # ═══════════════════════════════════════════════════════════════════
        story.append(PageBreak())
        fail_df = df[df['status'] == 'FAIL'].reset_index(drop=True) if 'status' in df.columns else pd.DataFrame()
        story.append(banner(f"❌  Failed / ATKT Students  ({len(fail_df)} students)"))
        story.append(Spacer(1, 0.1*inch))

        if fail_df.empty:
            story.append(Paragraph("🎉  No failed students — excellent result!", styles['Normal']))
        else:
            f_hdrs = ['Roll No', 'Student Name', 'Total', 'SGPA', 'Grade', 'ATKT', 'Failed Subjects']
            f_fixed = 0.6 + 1.7 + 0.5 + 0.5 + 0.45 + 0.5
            f_sub_w = CW - f_fixed * inch
            f_cws = [0.6*inch, 1.7*inch, 0.5*inch, 0.5*inch, 0.45*inch, 0.5*inch, f_sub_w]

            fail_data = [[Paragraph(h, hdr_style()) for h in f_hdrs]]
            for _, row in fail_df.iterrows():
                fsubs = str(row.get('failed_subjects', ''))
                cells = [str(row.get('roll_no', '')),
                         Paragraph(str(row.get('name', '')), body_style()),
                         str(row.get('total', '')), str(row.get('sgpa', '')),
                         str(row.get('grade', '')), str(row.get('atkt_count', 0)),
                         Paragraph(fsubs, body_style())]
                fail_data.append(cells)

            ts_fail = list(TBL_BASE)
            t3 = Table(fail_data, colWidths=f_cws, repeatRows=1)
            t3.setStyle(TableStyle(ts_fail))
            story.append(t3)

        # ═══════════════════════════════════════════════════════════════════
        # SECTION 4: Subject-wise Performance + Subject Toppers
        # ═══════════════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(banner("📊  Subject-wise Performance Summary"))
        story.append(Spacer(1, 0.1*inch))

        # Summary table (one row per subject)
        s_hdrs = ['Subject', 'Total', 'Pass', 'Fail', 'Avg', 'High', 'Low', 'Pass%', 'Topper', 'Marks']
        s_fixed = 0.5 + 0.5 + 0.5 + 0.5 + 0.5 + 0.5 + 0.5 + 1.2 + 0.5
        s_name_w = CW - s_fixed * inch
        s_cws = ([s_name_w, 0.5*inch, 0.5*inch, 0.5*inch,
                  0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 1.2*inch, 0.5*inch])

        sub_data = [[Paragraph(h, hdr_style()) for h in s_hdrs]]
        for i in range(7):
            col = f'sub_{i+1}'
            if col not in df.columns:
                continue
            sname = sub_names[i]
            marks = pd.to_numeric(df[col], errors='coerce')
            valid = marks[marks > 0]
            total_s = len(df)
            try:
                passing = final_result.EXT_PASSING_MIN[i] if PARSER_LOADED else 18
            except Exception:
                passing = 18
            if 'failed_subjects' in df.columns and df['failed_subjects'].astype(str).str.strip().ne('').any():
                fail_mask = df['failed_subjects'].astype(str).str.contains(sname, regex=False, na=False)
                fail_cnt = int(fail_mask.sum())
            else:
                fail_cnt = int((marks < passing).sum())
            pass_cnt = total_s - fail_cnt
            avg_m = f"{valid.mean():.1f}" if len(valid) else "—"
            high_m = str(int(valid.max())) if len(valid) else "—"
            low_m  = str(int(valid.min())) if len(valid) else "—"
            pass_pct = f"{pass_cnt/total_s*100:.1f}%" if total_s else "—"
            # Topper
            if len(valid):
                top_idx = marks.idxmax()
                top_name = str(df.loc[top_idx, 'name']) if top_idx in df.index else "—"
                top_mark = str(int(marks.loc[top_idx]))
            else:
                top_name, top_mark = "—", "—"

            sub_data.append([
                Paragraph(sname, body_style()),
                str(total_s), str(pass_cnt), str(fail_cnt),
                avg_m, high_m, low_m, pass_pct,
                Paragraph(top_name, body_style()), top_mark
            ])

        ts_sub = list(TBL_BASE)
        t4 = Table(sub_data, colWidths=s_cws, repeatRows=1)
        t4.setStyle(TableStyle(ts_sub))
        story.append(t4)

        # ═══════════════════════════════════════════════════════════════════
        # SECTION 5: Overall Result Summary  (Complete student list)
        # ═══════════════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(banner("📋  Overall Result Summary"))
        story.append(Spacer(1, 0.1 * inch))

        # ── KPI summary bar ─────────────────────────────────────────────────
        total_st   = len(df)
        pass_st    = int((df['status'] == 'PASS').sum()) if 'status' in df.columns else 0
        fail_st    = total_st - pass_st
        pass_pct_ov = f"{pass_st / total_st * 100:.1f}%" if total_st else "—"
        try:
            sgpa_num = pd.to_numeric(df['sgpa'], errors='coerce').dropna()
            sgpa_num = sgpa_num[sgpa_num > 0]
            avg_sgpa_ov = f"{sgpa_num.mean():.2f}" if len(sgpa_num) else "—"
            max_sgpa_ov = f"{sgpa_num.max():.2f}" if len(sgpa_num) else "—"
        except Exception:
            avg_sgpa_ov, max_sgpa_ov = "—", "—"

        kpi_ps = ParagraphStyle('kpi_lbl', fontName='Helvetica', fontSize=7.5,
                                textColor=colors.HexColor('#64748b'))
        kpi_val_ps = ParagraphStyle('kpi_val', fontName='Helvetica-Bold', fontSize=14,
                                    textColor=colors.HexColor('#0c1a3a'))
        kpi_green  = ParagraphStyle('kpi_g',  fontName='Helvetica-Bold', fontSize=14,
                                    textColor=colors.HexColor('#16a34a'))
        kpi_red    = ParagraphStyle('kpi_r',  fontName='Helvetica-Bold', fontSize=14,
                                    textColor=colors.HexColor('#dc2626'))
        kpi_purple = ParagraphStyle('kpi_p',  fontName='Helvetica-Bold', fontSize=14,
                                    textColor=colors.HexColor('#7c3aed'))

        def kpi_cell(label, value, val_style):
            return [Paragraph(value, val_style), Paragraph(label, kpi_ps)]

        kpi_col_w = CW / 5
        kpi_data = [[
            kpi_cell("Total Students",  str(total_st),    kpi_val_ps),
            kpi_cell("Passed",          str(pass_st),     kpi_green),
            kpi_cell("Failed",          str(fail_st),     kpi_red),
            kpi_cell("Pass %",          pass_pct_ov,      kpi_green),
            kpi_cell("Avg SGPA",        avg_sgpa_ov,      kpi_purple),
        ]]
        kpi_tbl = Table(kpi_data, colWidths=[kpi_col_w] * 5)
        kpi_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), colors.HexColor('#f0f9ff')),
            ('BOX',           (0, 0), (-1, -1), 0.8, colors.HexColor('#0c1a3a')),
            ('INNERGRID',     (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING',    (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(kpi_tbl)
        story.append(Spacer(1, 0.15 * inch))

        # ── Grade distribution summary ───────────────────────────────────────
        grade_counts = {}
        if 'grade' in df.columns:
            for g in ['O', 'A+', 'A', 'B+', 'B', 'C', 'F']:
                cnt = int((df['grade'] == g).sum())
                grade_counts[g] = cnt

        if grade_counts:
            gd_ps  = ParagraphStyle('gd', fontName='Helvetica-Bold', fontSize=8,
                                    textColor=colors.white)
            gd_ps2 = ParagraphStyle('gd2', fontName='Helvetica', fontSize=8,
                                    textColor=colors.HexColor('#1e293b'))
            grade_colors_map = {
                'O': '#0d9488', 'A+': '#0284c7', 'A': '#2563eb',
                'B+': '#7c3aed', 'B': '#d97706', 'C': '#64748b', 'F': '#dc2626'
            }
            g_col_w = CW / len(grade_counts)
            g_data = [[Paragraph(g, gd_ps) for g in grade_counts.keys()],
                      [str(v) for v in grade_counts.values()]]
            g_tbl = Table(g_data, colWidths=[g_col_w] * len(grade_counts))
            g_style = [
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING',    (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('FONTNAME',      (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 1), (-1, 1), 11),
                ('BOX',           (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('INNERGRID',     (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
            ]
            for ci, g in enumerate(grade_counts.keys()):
                hex_c = grade_colors_map.get(g, '#0c1a3a')
                g_style.append(('BACKGROUND', (ci, 0), (ci, 0), colors.HexColor(hex_c)))
                txt_c = colors.HexColor('#dc2626') if g == 'F' else colors.HexColor('#0c1a3a')
                g_style.append(('TEXTCOLOR', (ci, 1), (ci, 1), txt_c))
            g_tbl.setStyle(TableStyle(g_style))
            story.append(Paragraph("Grade Distribution", ParagraphStyle(
                'gd_title', fontName='Helvetica-Bold', fontSize=9,
                textColor=colors.HexColor('#0c1a3a'))))
            story.append(Spacer(1, 0.05 * inch))
            story.append(g_tbl)
            story.append(Spacer(1, 0.15 * inch))

        # ── Complete Student Result Table ────────────────────────────────────
        story.append(Paragraph("Complete Student Result", ParagraphStyle(
            'cr_title', fontName='Helvetica-Bold', fontSize=9,
            textColor=colors.HexColor('#0c1a3a'))))
        story.append(Spacer(1, 0.05 * inch))

        all_hdrs = ['#', 'Roll No', 'Student Name', 'Total', 'SGPA', 'Grade', 'Status', 'ATKT']
        fixed_w  = 0.3 + 0.6 + 0.55 + 0.55 + 0.55 + 0.65 + 0.45
        name_w   = CW - fixed_w * inch
        all_cws  = [0.3*inch, 0.6*inch, name_w,
                    0.55*inch, 0.55*inch, 0.55*inch, 0.65*inch, 0.45*inch]

        all_data = [[Paragraph(h, hdr_style()) for h in all_hdrs]]
        ts_all   = list(TBL_BASE)

        sorted_df = df.sort_values('roll_no').reset_index(drop=True) if 'roll_no' in df.columns else df.reset_index(drop=True)
        for idx, (_, row) in enumerate(sorted_df.iterrows(), start=1):
            st = str(row.get('status', ''))
            cells = [
                str(idx),
                str(row.get('roll_no', '')),
                Paragraph(str(row.get('name', '')), body_style()),
                str(row.get('total', '')),
                str(row.get('sgpa', '')),
                str(row.get('grade', '')),
                st,
                str(row.get('atkt_count', 0)),
            ]
            all_data.append(cells)
            st_col = all_hdrs.index('Status')
            col_c  = (colors.HexColor('#16a34a') if st == 'PASS'
                      else colors.HexColor('#dc2626'))
            ts_all.append(('TEXTCOLOR', (st_col, idx), (st_col, idx), col_c))
            if st == 'FAIL':
                ts_all.append(('FONTNAME', (st_col, idx), (st_col, idx), 'Helvetica-Bold'))
                ts_all.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#fff1f2')))

        t5 = Table(all_data, colWidths=all_cws, repeatRows=1)
        t5.setStyle(TableStyle(ts_all))
        story.append(t5)

        doc.build(story)


    def export_to_pdf(self, data, filename="student_report.pdf"):
        """Simple quick-export PDF (called from dashboard export button)."""
        if data is None or (hasattr(data, 'empty') and data.empty):
            QMessageBox.warning(self, "No Data", "No data available to export.")
            return
        _default_pdf2 = os.path.join(BASE_DIR, filename)
        path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", _default_pdf2, "PDF Files (*.pdf)")
        if not path:
            return
        if not path.lower().endswith(".pdf"):
            path += ".pdf"
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        try:
            self.generate_comprehensive_pdf(data, path, section='all')
            QMessageBox.information(self, "✅ Success", f"PDF exported to:\n{path}")
        except PermissionError:
            QMessageBox.critical(self, "File Locked 🔒",
                f"Cannot save PDF — the file is already open in another program.\n\n"
                f"❌ Close it in your PDF viewer first, then try again:\n{path}")
        except Exception as e:
            import traceback; traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Failed to export PDF:\n{e}")


    def generate_report_html(self, df):
        """Generate simple HTML report for printing."""
        total = len(df)
        pass_c = len(df[df['status'] == 'PASS']) if 'status' in df.columns else 0
        fail_c = total - pass_c
        date_str = QDate.currentDate().toString("dd MMM yyyy")
        rows_html = ""
        for _, row in df.iterrows():
            color = "#16a34a" if str(row.get('status','')) == 'PASS' else "#dc2626"
            rows_html += f"""<tr>
                <td>{row.get('roll_no','')}</td>
                <td>{row.get('name','')}</td>
                <td>{row.get('total','')}</td>
                <td>{row.get('sgpa','')}</td>
                <td>{row.get('grade','')}</td>
                <td style="color:{color};font-weight:700">{row.get('status','')}</td>
            </tr>"""
        return f"""<!DOCTYPE html><html><head>
        <style>
            body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; color: #1e293b; }}
            h1 {{ color: #0c1a3a; font-size: 18px; margin-bottom: 4px; }}
            p {{ color: #64748b; margin: 0 0 12px 0; }}
            .summary {{ display: flex; gap: 16px; margin-bottom: 16px; }}
            .stat-box {{ background: #f0f4f8; padding: 8px 16px; border-radius: 6px; }}
            .stat-box b {{ font-size: 20px; color: #0d9488; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background: #0c1a3a; color: white; padding: 8px 10px; text-align: left; font-size: 11px; }}
            td {{ padding: 6px 10px; border-bottom: 1px solid #e2e8f0; }}
            tr:nth-child(even) {{ background: #f8fafc; }}
        </style></head><body>
        <h1>Student Result Report — VNSGU SASCMA</h1>
        <p>Generated: {date_str}</p>
        <div class="summary">
            <div class="stat-box"><b>{total}</b><br>Total Students</div>
            <div class="stat-box"><b style="color:#16a34a">{pass_c}</b><br>Passed</div>
            <div class="stat-box"><b style="color:#dc2626">{fail_c}</b><br>Failed</div>
        </div>
        <table>
            <thead><tr>
                <th>Roll No</th><th>Name</th><th>Total</th>
                <th>SGPA</th><th>Grade</th><th>Status</th>
            </tr></thead>
            <tbody>{rows_html}</tbody>
        </table></body></html>"""

    def print_report(self):
        """Generate comprehensive PDF and open in system PDF viewer (preview + native print, Ctrl+P)."""
        if not self.data_loaded:
            QMessageBox.warning(self, "No Data", "Please upload a PDF first.")
            return
        df = self.db.get_all_data()
        if df.empty:
            QMessageBox.warning(self, "No Data", "No data available to print.")
            return
        try:
            import tempfile
            # Generate comprehensive PDF to a temp file (same as Generate Report → All)
            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False, prefix='student_report_')
            tmp_path = tmp.name
            tmp.close()
            self.generate_comprehensive_pdf(df, tmp_path, 'all')
            # Open with system default PDF viewer — user gets preview + native print dialog
            os.startfile(tmp_path)
        except Exception as e:
            import traceback; traceback.print_exc()
            QMessageBox.critical(self, "Print Error",
                                 f"Could not open report for printing:\n{e}\n\nTry: Generate Report → Save PDF, then print from PDF viewer.")

    def generate_print_html(self, df):
        """Generate comprehensive HTML for print preview (mirrors PDF report structure)."""
        date_str = QDate.currentDate().toString("dd MMMM yyyy")
        total = len(df)
        pass_c = len(df[df['status'] == 'PASS']) if 'status' in df.columns else 0
        fail_c = total - pass_c

        sub_names = [self.get_subject_name(i) for i in range(7)]

        # ── Section 1: Top Rankers (sorted by SGPA) ────────────────────────
        try:
            top_df = self.sort_by_sgpa_then_roll_no(df).reset_index(drop=True)
        except Exception:
            top_df = df.copy()

        rank_rows = ""
        rank_cls = ['gold', 'silver', 'bronze']
        rank_numbers = self.sgpa_rank_numbers(top_df)
        for (_, row), rank_no in zip(top_df.iterrows(), rank_numbers):
            cls = rank_cls[rank_no-1] if rank_no <= 3 else ""
            medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(rank_no, str(rank_no))
            st = str(row.get('status', ''))
            sc = 'pass' if st == 'PASS' else 'fail'
            rank_rows += (f'<tr class="{cls}"><td>{medal}</td>'
                         f'<td>{row.get("roll_no","")}</td>'
                         f'<td>{row.get("name","")}</td>'
                         f'<td>{row.get("total","")}</td>'
                         f'<td>{row.get("sgpa","")}</td>'
                         f'<td>{row.get("grade","")}</td>'
                         f'<td class="{sc}">{st}</td></tr>')

        # ── Section 2: Failed Students ──────────────────────────────────────
        fail_df = df[df['status'] == 'FAIL'] if 'status' in df.columns else df.iloc[0:0]
        fail_rows = ""
        for _, row in fail_df.iterrows():
            fail_rows += (f'<tr><td>{row.get("roll_no","")}</td>'
                         f'<td>{row.get("name","")}</td>'
                         f'<td>{row.get("total","")}</td>'
                         f'<td>{row.get("sgpa","")}</td>'
                         f'<td>{row.get("grade","")}</td>'
                         f'<td>{row.get("atkt_count",0)}</td>'
                         f'<td>{str(row.get("failed_subjects","")).replace(chr(10),", ")}</td></tr>')
        if not fail_rows:
            fail_rows = '<tr><td colspan="7" style="text-align:center;color:#16a34a;">🎉 No failed students</td></tr>'

        # ── Section 3: Subject-wise Summary ────────────────────────────────
        sub_rows = ""
        for i in range(7):
            col = f'sub_{i+1}'
            if col not in df.columns:
                continue
            sname = sub_names[i]
            marks = pd.to_numeric(df[col], errors='coerce')
            valid = marks[marks > 0]
            total_s = len(df)
            try:
                passing = final_result.EXT_PASSING_MIN[i] if PARSER_LOADED else 18
            except Exception:
                passing = 18
            if 'failed_subjects' in df.columns:
                fail_cnt = int(df['failed_subjects'].astype(str).str.contains(sname, regex=False, na=False).sum())
            else:
                fail_cnt = int((marks < passing).sum())
            pass_cnt = total_s - fail_cnt
            avg_m = f"{valid.mean():.1f}" if len(valid) else "—"
            high_m = str(int(valid.max())) if len(valid) else "—"
            low_m  = str(int(valid.min())) if len(valid) else "—"
            pass_pct = f"{pass_cnt/total_s*100:.1f}%" if total_s else "—"
            if len(valid):
                top_idx = marks.idxmax()
                top_name = str(df.loc[top_idx, 'name']) if top_idx in df.index else "—"
                top_mark = str(int(marks.loc[top_idx]))
            else:
                top_name, top_mark = "—", "—"
            sub_rows += (f'<tr><td>{sname}</td><td>{total_s}</td><td>{pass_cnt}</td>'
                        f'<td>{fail_cnt}</td><td>{avg_m}</td><td>{high_m}</td>'
                        f'<td>{low_m}</td><td>{pass_pct}</td>'
                        f'<td>{top_name}</td><td>{top_mark}</td></tr>')

        return f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head><body>
<p style="text-align:center;font-size:8pt;color:#64748b;margin:0 0 6px 0;">
    {COURSE_NAME} &nbsp;&bull;&nbsp; Generated: {date_str} &nbsp;&bull;&nbsp;
    Total: {total} &nbsp;|&nbsp; Pass: {pass_c} &nbsp;|&nbsp; Fail: {fail_c}
</p>

<h2>&#127942; Top Rankers (Sorted by SGPA)</h2>
<table>
<thead><tr><th>Rank</th><th>Roll No</th><th>Name</th><th>Total</th>
<th>SGPA</th><th>Grade</th><th>Status</th></tr></thead>
<tbody>{rank_rows}</tbody>
</table>

<div style="page-break-before:always;"></div>
<h2>&#10060; Failed / ATKT Students ({fail_c} students)</h2>
<table>
<thead><tr><th>Roll No</th><th>Name</th><th>Total</th><th>SGPA</th>
<th>Grade</th><th>ATKT</th><th>Failed Subjects</th></tr></thead>
<tbody>{fail_rows}</tbody>
</table>

<div style="page-break-before:always;"></div>
<h2>&#128202; Subject-wise Performance Summary</h2>
<table>
<thead><tr><th>Subject</th><th>Total</th><th>Pass</th><th>Fail</th>
<th>Avg</th><th>High</th><th>Low</th><th>Pass%</th>
<th>Topper</th><th>Marks</th></tr></thead>
<tbody>{sub_rows}</tbody>
</table>

<p class="footer">Printed by Student Result Analyzer &mdash; {COURSE_NAME}</p>
</body></html>"""

    # ──────────────────────────────────────────────────────────────────────────
    # COLLEGE WISE REPORT PAGE
    # ──────────────────────────────────────────────────────────────────────────
    def create_college_wise_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        # Title
        pg_title = QLabel("🏫  College Wise Report")
        pg_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0c1a3a;")
        layout.addWidget(pg_title)

        # ── Controls card ──────────────────────────────────────────────────
        ctrl_card = QFrame()
        ctrl_card.setObjectName("card")
        ctrl_card.setFixedHeight(70)
        ctrl_layout = QHBoxLayout(ctrl_card)
        ctrl_layout.setContentsMargins(18, 10, 18, 10)
        ctrl_layout.setSpacing(12)

        ctrl_layout.addWidget(QLabel("Select College:"))

        self.college_combo = QComboBox()
        self.college_combo.setFixedHeight(36)
        self.college_combo.setMinimumWidth(320)
        self.college_combo.setCursor(QCursor(Qt.PointingHandCursor))
        self.college_combo.currentIndexChanged.connect(self._on_college_changed)
        ctrl_layout.addWidget(self.college_combo)

        btn_refresh_college = QPushButton("🔄  Refresh")
        btn_refresh_college.setFixedHeight(36)
        btn_refresh_college.setObjectName("secondary")
        btn_refresh_college.setCursor(QCursor(Qt.PointingHandCursor))
        btn_refresh_college.clicked.connect(self.refresh_college_wise)
        ctrl_layout.addWidget(btn_refresh_college)

        ctrl_layout.addStretch()

        btn_excel_college = QPushButton("📥  Export Excel")
        btn_excel_college.setFixedHeight(36)
        btn_excel_college.setObjectName("secondary")
        btn_excel_college.setCursor(QCursor(Qt.PointingHandCursor))
        btn_excel_college.clicked.connect(self._export_college_excel)
        ctrl_layout.addWidget(btn_excel_college)

        btn_pdf_college = QPushButton("📄  Export PDF")
        btn_pdf_college.setFixedHeight(36)
        btn_pdf_college.setCursor(QCursor(Qt.PointingHandCursor))
        btn_pdf_college.clicked.connect(self._export_college_pdf)
        ctrl_layout.addWidget(btn_pdf_college)

        layout.addWidget(ctrl_card)

        # ── KPI row ───────────────────────────────────────────────────────
        self.college_kpi_row = QHBoxLayout()
        self.college_kpi_row.setSpacing(12)
        self._college_kpi_total   = KPICard("Total Students",  "—", "#0c1a3a", "👥")
        self._college_kpi_pass    = KPICard("Passed",          "—", "#16a34a", "✅")
        self._college_kpi_fail    = KPICard("Failed / ATKT",   "—", "#dc2626", "❌")
        self._college_kpi_passperc= KPICard("Pass %",          "—", "#0891b2", "📊")
        self._college_kpi_sgpa    = KPICard("Avg SGPA",        "—", "#7c3aed", "📈")
        for kpi in [self._college_kpi_total, self._college_kpi_pass,
                    self._college_kpi_fail,  self._college_kpi_passperc,
                    self._college_kpi_sgpa]:
            self.college_kpi_row.addWidget(kpi)
        layout.addLayout(self.college_kpi_row)

        # ── Chart row (visible only in All-Colleges mode) ──────────────────
        self.college_chart_outer = QFrame()
        self.college_chart_outer.setObjectName("card")
        chart_row_layout = QHBoxLayout(self.college_chart_outer)
        chart_row_layout.setContentsMargins(12, 12, 12, 12)
        chart_row_layout.setSpacing(14)

        # Bar chart — Pass% per college
        self.col_bar_canvas = MplCanvas(self, width=5, height=3.6, dpi=90)
        chart_row_layout.addWidget(self.col_bar_canvas, 3)

        # Pie chart — overall pass/fail among all students
        self.col_pie_canvas = MplCanvas(self, width=3.2, height=3.6, dpi=90)
        chart_row_layout.addWidget(self.col_pie_canvas, 2)

        self.college_chart_outer.setVisible(False)
        layout.addWidget(self.college_chart_outer)

        # ── Overview table (All-Colleges summary) ─────────────────────────
        self.col_overview_outer, _, _, col_overview_inner = make_section_card(
            "All Colleges — Overview", collapsible=False
        )
        self.college_overview_table = QTableWidget()
        self.college_overview_table.setColumnCount(7)
        self.college_overview_table.setHorizontalHeaderLabels(
            ["College Name", "Total", "Pass", "Fail", "Pass %", "Avg SGPA", "ATKT"])
        self.setup_table(self.college_overview_table, row_height=34)
        self.college_overview_table.setMinimumHeight(200)
        ov_hdr = self.college_overview_table.horizontalHeader()
        ov_hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        for ci in range(1, 7):
            ov_hdr.setSectionResizeMode(ci, QHeaderView.ResizeToContents)
        col_overview_inner.addWidget(self.college_overview_table)
        self.col_overview_outer.setVisible(False)
        layout.addWidget(self.col_overview_outer)

        # ── Table ─────────────────────────────────────────────────────────
        tbl_outer, _, _, tbl_inner = make_section_card("📋  Students", collapsible=False)

        self.college_table = QTableWidget()
        self.college_table.setColumnCount(7)
        self.college_table.setHorizontalHeaderLabels(
            ["Roll No", "Name", "SGPA", "Total Marks", "Grade", "Status", "ATKT Count"])
        self.setup_table(self.college_table, row_height=34)
        self.college_table.setMinimumHeight(380)

        hdr = self.college_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(6, QHeaderView.ResizeToContents)

        tbl_inner.addWidget(self.college_table)
        layout.addWidget(tbl_outer)

        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def refresh_college_wise(self):
        """Populate college dropdown from DB data."""
        if not self.data_loaded:
            return
        df = self.db.get_all_data()
        if df.empty:
            return

        colleges = sorted(
            [c for c in df['college'].dropna().unique() if str(c).strip()],
            key=lambda x: str(x).strip()
        )

        self.college_combo.blockSignals(True)
        prev = self.college_combo.currentText()
        self.college_combo.clear()

        # First: All Colleges overview option
        self.college_combo.addItem("All Colleges (Overview)", "__ALL__")
        # Separator
        self.college_combo.insertSeparator(1)
        # Then individual colleges
        for col in colleges:
            self.college_combo.addItem(str(col).strip(), str(col).strip())

        # Restore previous selection if still available
        idx = self.college_combo.findText(prev)
        if idx >= 0:
            self.college_combo.setCurrentIndex(idx)
        else:
            self.college_combo.setCurrentIndex(0)   # default: All Colleges
        self.college_combo.blockSignals(False)

        self._on_college_changed()

    def _on_college_changed(self):
        """Filter table and update KPIs for selected college, or show All-Colleges overview."""
        college = self.college_combo.currentData()

        if not self.data_loaded:
            self.college_table.setRowCount(0)
            for kpi in [self._college_kpi_total, self._college_kpi_pass,
                        self._college_kpi_fail, self._college_kpi_passperc,
                        self._college_kpi_sgpa]:
                kpi.lbl_value.setText("—")
            return

        df = self.db.get_all_data()

        # ── ALL COLLEGES MODE ──────────────────────────────────────────────
        if college == "__ALL__":
            self.college_chart_outer.setVisible(True)
            self.col_overview_outer.setVisible(True)

            # KPIs — totals across all students
            total  = len(df)
            passed = int((df['status'] == 'PASS').sum())
            failed = total - passed
            pass_perc = f"{(passed/total*100):.1f}%" if total else "0%"
            sgpa_vals = pd.to_numeric(df['sgpa'], errors='coerce').dropna()
            sgpa_vals = sgpa_vals[sgpa_vals > 0]
            avg_sgpa  = f"{sgpa_vals.mean():.2f}" if len(sgpa_vals) else "—"

            self._college_kpi_total.lbl_value.setText(str(total))
            self._college_kpi_pass.lbl_value.setText(str(passed))
            self._college_kpi_fail.lbl_value.setText(str(failed))
            self._college_kpi_passperc.lbl_value.setText(pass_perc)
            self._college_kpi_sgpa.lbl_value.setText(avg_sgpa)

            # Build per-college summary using DB method
            college_stats = self.db.get_college_stats()

            # ── Overview table ──
            if not college_stats.empty:
                self.college_overview_table.setRowCount(len(college_stats))
                palette = ["#f0fdf4", "#fef2f2", "#eff6ff", "#fefce8",
                           "#fdf4ff", "#fff7ed", "#f0fdfa"]
                for r_idx, (_, row) in enumerate(college_stats.iterrows()):
                    tot_r  = int(row.get('total_students', 0))
                    pass_r = int(row.get('pass_count', 0))
                    fail_r = int(row.get('fail_count', 0))
                    pp_r   = row.get('pass_percentage', 0)
                    sg_r   = row.get('avg_sgpa', None)
                    # Count ATKT students per college
                    col_name = str(row.get('college', ''))
                    col_df   = df[df['college'].astype(str).str.strip() == col_name]
                    atkt_r   = int((col_df['atkt_count'] > 0).sum())

                    row_vals = [
                        col_name,
                        str(tot_r),
                        str(pass_r),
                        str(fail_r),
                        f"{pp_r:.1f}%",
                        f"{sg_r:.2f}" if sg_r and sg_r == sg_r else "—",
                        str(atkt_r),
                    ]
                    bg = QColor(palette[r_idx % len(palette)])
                    for c_idx, val in enumerate(row_vals):
                        item = QTableWidgetItem(val)
                        item.setTextAlignment(
                            Qt.AlignLeft | Qt.AlignVCenter if c_idx == 0
                            else Qt.AlignCenter
                        )
                        item.setBackground(bg)
                        # Color pass% column green/red
                        if c_idx == 4:
                            try:
                                pp_val = float(str(pp_r))
                                item.setForeground(
                                    QColor("#16a34a") if pp_val >= 50 else QColor("#dc2626")
                                )
                            except Exception:
                                pass
                        self.college_overview_table.setItem(r_idx, c_idx, item)
            else:
                self.college_overview_table.setRowCount(0)

            # ── Charts ──
            try:
                tc = '#0c1a3a'
                if not college_stats.empty:
                    names  = [str(n)[:18] for n in college_stats['college']]
                    ppercs = college_stats['pass_percentage'].fillna(0).tolist()
                    bar_colors = [
                        '#16a34a' if p >= 50 else '#dc2626' for p in ppercs
                    ]

                    # Bar chart: college vs pass%
                    self.col_bar_canvas.fig.patch.set_facecolor('#ffffff')
                    self.col_bar_canvas.axes.clear()
                    self.col_bar_canvas.axes.set_facecolor('#f8fafc')
                    bars = self.col_bar_canvas.axes.barh(
                        names, ppercs, color=bar_colors,
                        edgecolor='#ffffff', linewidth=1.2
                    )
                    # Value labels
                    for bar, val in zip(bars, ppercs):
                        self.col_bar_canvas.axes.text(
                            min(val + 1, 102), bar.get_y() + bar.get_height()/2,
                            f"{val:.1f}%", va='center', ha='left',
                            fontsize=8, color=tc, fontweight='bold'
                        )
                    self.col_bar_canvas.axes.set_xlim(0, 115)
                    self.col_bar_canvas.axes.set_xlabel(
                        "Pass Percentage", color='#64748b', fontsize=9
                    )
                    self.col_bar_canvas.axes.set_title(
                        "College-wise Pass %", color=tc,
                        fontsize=11, fontweight='bold'
                    )
                    self.col_bar_canvas.axes.axvline(
                        x=50, color='#f97316', linewidth=1.2,
                        linestyle='--', label='50% line'
                    )
                    self.col_bar_canvas.axes.tick_params(colors='#64748b', labelsize=8)
                    self.col_bar_canvas.axes.spines['top'].set_visible(False)
                    self.col_bar_canvas.axes.spines['right'].set_visible(False)
                    self.col_bar_canvas.axes.spines['left'].set_color('#e2e8f0')
                    self.col_bar_canvas.axes.spines['bottom'].set_color('#e2e8f0')
                    self.col_bar_canvas.axes.grid(
                        axis='x', color='#e2e8f0', linewidth=0.7
                    )
                    self.col_bar_canvas.fig.tight_layout()
                    self.col_bar_canvas.draw()

                    # Pie chart: total pass vs fail
                    self.col_pie_canvas.fig.patch.set_facecolor('#ffffff')
                    self.col_pie_canvas.axes.clear()
                    self.col_pie_canvas.axes.set_facecolor('#ffffff')
                    pass_t = college_stats['pass_count'].sum()
                    fail_t = college_stats['fail_count'].sum()
                    if pass_t + fail_t > 0:
                        self.col_pie_canvas.axes.pie(
                            [pass_t, fail_t],
                            labels=[f"Pass ({int(pass_t)})", f"Fail ({int(fail_t)})"],
                            autopct='%1.1f%%',
                            startangle=90,
                            colors=['#0d9488', '#f97316'],
                            wedgeprops={'linewidth': 2.5, 'edgecolor': '#ffffff'},
                            textprops={'color': tc, 'fontsize': 9, 'fontweight': 'bold'}
                        )
                        self.col_pie_canvas.axes.set_title(
                            "Overall Pass / Fail", color=tc,
                            fontsize=11, fontweight='bold'
                        )
                    self.col_pie_canvas.fig.tight_layout()
                    self.col_pie_canvas.draw()
            except Exception as e:
                print(f"[ColChart] {e}")

            # Student table — all students sorted by college
            all_sorted = df.sort_values(
                by=['college', 'status'], ascending=[True, True]
            ).reset_index(drop=True)
            self._populate_college_student_table(all_sorted)
            self._college_filtered_df = all_sorted.copy()
            return

        # ── SINGLE COLLEGE MODE ────────────────────────────────────────────
        self.college_chart_outer.setVisible(False)
        self.col_overview_outer.setVisible(False)

        if not college:
            self.college_table.setRowCount(0)
            for kpi in [self._college_kpi_total, self._college_kpi_pass,
                        self._college_kpi_fail, self._college_kpi_passperc,
                        self._college_kpi_sgpa]:
                kpi.lbl_value.setText("—")
            return

        filtered = df[df['college'].astype(str).str.strip() == college].copy()
        filtered = filtered.reset_index(drop=True)

        total  = len(filtered)
        passed = int((filtered['status'] == 'PASS').sum())
        failed = total - passed
        pass_perc = f"{(passed/total*100):.1f}%" if total else "0%"
        sgpa_vals = pd.to_numeric(filtered['sgpa'], errors='coerce').dropna()
        sgpa_vals = sgpa_vals[sgpa_vals > 0]
        avg_sgpa  = f"{sgpa_vals.mean():.2f}" if len(sgpa_vals) else "—"

        self._college_kpi_total.lbl_value.setText(str(total))
        self._college_kpi_pass.lbl_value.setText(str(passed))
        self._college_kpi_fail.lbl_value.setText(str(failed))
        self._college_kpi_passperc.lbl_value.setText(pass_perc)
        self._college_kpi_sgpa.lbl_value.setText(avg_sgpa)

        try:
            filtered = self.sort_by_sgpa_then_roll_no(filtered)
        except Exception:
            pass

        self._populate_college_student_table(filtered)
        self._college_filtered_df = filtered.copy()

    def _populate_college_student_table(self, df_in):
        """Fill college_table with rows from df_in."""
        self.college_table.setRowCount(len(df_in))
        for r_idx, (_, row) in enumerate(df_in.iterrows()):
            status = str(row.get('status', ''))
            items = [
                str(row.get('roll_no', '')),
                str(row.get('name', '')),
                str(row.get('sgpa', '')),
                str(row.get('total', '')),
                str(row.get('grade', '')),
                status,
                str(row.get('atkt_count', 0)),
            ]
            for c_idx, val in enumerate(items):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                if c_idx == 5:  # Status column
                    if status == 'PASS':
                        item.setForeground(QColor("#16a34a"))
                        item.setBackground(QColor("#f0fdf4"))
                    else:
                        item.setForeground(QColor("#dc2626"))
                        item.setBackground(QColor("#fef2f2"))
                self.college_table.setItem(r_idx, c_idx, item)

    def _export_college_excel(self):
        college = self.college_combo.currentData()
        if not college:
            QMessageBox.warning(self, "No College", "Please select a college first.")
            return
        df = getattr(self, '_college_filtered_df', None)
        if df is None or df.empty:
            QMessageBox.warning(self, "No Data", "No data to export.")
            return
        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in college)
        self.export_to_excel(df, f"College_{safe_name}_Report.xlsx")

    def _export_college_pdf(self):
        college = self.college_combo.currentData()
        if not college:
            QMessageBox.warning(self, "No College", "Please select a college first.")
            return
        df = getattr(self, '_college_filtered_df', None)
        if df is None or df.empty:
            QMessageBox.warning(self, "No Data", "No data to export.")
            return
        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in college)
        default_path = os.path.join(BASE_DIR, f"College_{safe_name}_Report.pdf")
        path, _ = QFileDialog.getSaveFileName(self, "Save College PDF", default_path, "PDF Files (*.pdf)")
        if not path:
            return
        if not path.lower().endswith(".pdf"):
            path += ".pdf"
        try:
            self.generate_comprehensive_pdf(df, path, section='all')
            QMessageBox.information(self, "✅ Saved", f"PDF saved:\n{path}")
        except PermissionError:
            QMessageBox.critical(self, "File Locked 🔒",
                f"File is open in another program.\nClose it and try again:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "PDF Error", f"Could not generate PDF:\n{e}")

    def create_settings_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        pg_title = QLabel("⚙  Settings & Information")
        pg_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0c1a3a;")
        layout.addWidget(pg_title)

        # Status Card
        status_outer, _, _, status_inner = make_section_card("📊  Data Status", collapsible=False)
        
        stats_row = QHBoxLayout()
        stats_row.setSpacing(14)
        
        def make_info_card(label, attr, color="#0c1a3a"):
            frm = QFrame()
            frm.setObjectName("kpi_card")
            frm.setStyleSheet(f"QFrame#kpi_card {{ border-top: 4px solid {color}; }}")
            fl = QVBoxLayout(frm)
            fl.setContentsMargins(14, 10, 14, 10)
            val = QLabel("—")
            val.setStyleSheet(f"font-size: 24px; font-weight: 800; color: {color};")
            lbl = QLabel(label)
            lbl.setStyleSheet("font-size: 11px; color: #64748b; font-weight: 600;")
            fl.addWidget(val)
            fl.addWidget(lbl)
            setattr(self, attr, val)
            return frm
        
        stats_row.addWidget(make_info_card("Total Students", "settings_total_lbl", "#0c1a3a"))
        stats_row.addWidget(make_info_card("Passed", "settings_pass_lbl", "#0d9488"))
        stats_row.addWidget(make_info_card("Failed", "settings_fail_lbl", "#dc2626"))
        stats_row.addWidget(make_info_card("Parser Status", "settings_parser_lbl", "#f97316"))
        status_inner.addLayout(stats_row)
        layout.addWidget(status_outer)

        # Quick Actions Card
        actions_outer, _, _, actions_inner = make_section_card("⚡  Quick Actions", collapsible=False)

        action_row_1 = QHBoxLayout()
        upload_btn = QPushButton("📤  Upload New PDF")
        upload_btn.setCursor(QCursor(Qt.PointingHandCursor))
        upload_btn.clicked.connect(lambda: self.switch_page(0))

        dashboard_btn = QPushButton("📊  Open Dashboard")
        dashboard_btn.setObjectName("secondary")
        dashboard_btn.setCursor(QCursor(Qt.PointingHandCursor))
        dashboard_btn.clicked.connect(lambda: self.switch_page(1))

        action_row_1.addWidget(upload_btn)
        action_row_1.addWidget(dashboard_btn)
        action_row_1.addStretch()

        action_row_2 = QHBoxLayout()

        generate_btn = QPushButton("📥  Generate Report")
        generate_btn.setCursor(QCursor(Qt.PointingHandCursor))
        generate_btn.setToolTip("Select report type (All/Pass/Fail) and format (Excel/PDF)")
        generate_btn.clicked.connect(self.show_report_dialog)

        print_btn = QPushButton("🖨  Print Report")
        print_btn.setObjectName("secondary")
        print_btn.setCursor(QCursor(Qt.PointingHandCursor))
        print_btn.setToolTip("Send result report to printer  [Ctrl+P]")
        print_btn.clicked.connect(self.print_report)

        export_failed_btn = QPushButton("❌  Export Failed Only")
        export_failed_btn.setObjectName("secondary")
        export_failed_btn.setCursor(QCursor(Qt.PointingHandCursor))
        export_failed_btn.clicked.connect(self.export_failed_students)

        action_row_2.addWidget(generate_btn)
        action_row_2.addWidget(print_btn)
        action_row_2.addWidget(export_failed_btn)
        action_row_2.addStretch()

        actions_inner.addLayout(action_row_1)
        actions_inner.addSpacing(8)
        actions_inner.addLayout(action_row_2)
        layout.addWidget(actions_outer)

        # Reset / Danger Zone
        danger_outer, _, _, danger_inner = make_section_card("⚠  Reset Application", collapsible=False)
        danger_inner.parent().setStyleSheet("QFrame#section_content_card { background: #fff7f7; border-radius: 0px 0px 6px 6px; }")
        
        danger_desc = QLabel("Clear all in-memory result data and lock analysis pages until a new PDF is uploaded.")
        danger_desc.setWordWrap(True)
        danger_desc.setStyleSheet("color: #64748b; font-size: 13px;")
        
        reset_btn = QPushButton("🗑  Reset Application")
        reset_btn.setObjectName("danger")
        reset_btn.setCursor(QCursor(Qt.PointingHandCursor))
        reset_btn.setFixedWidth(220)
        reset_btn.clicked.connect(self.clear_db)
        
        danger_inner.addWidget(danger_desc)
        danger_inner.addSpacing(10)
        danger_inner.addWidget(reset_btn)
        layout.addWidget(danger_outer)
        layout.addStretch()

        scroll.setWidget(content)
        page_layout.addWidget(scroll)
        return page

    def refresh_settings_page(self):
        if not hasattr(self, 'settings_total_lbl'):
            return
        df = self.db.get_all_data()
        total = len(df)
        pass_count = len(df[df['status'] == 'PASS']) if not df.empty and 'status' in df.columns else 0
        fail_count = len(df[df['status'] == 'FAIL']) if not df.empty and 'status' in df.columns else 0
        self.settings_total_lbl.setText(str(total))
        self.settings_pass_lbl.setText(str(pass_count))
        self.settings_fail_lbl.setText(str(fail_count))
        self.settings_parser_lbl.setText("Ready" if PARSER_LOADED else "Missing")

    # --- Logic Methods ---
    def filter_table(self, text):
        search_text = text.lower()
        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 1)
            roll_item = self.table.item(row, 0)
            
            match = False
            if name_item and search_text in name_item.text().lower():
                match = True
            elif roll_item and search_text in roll_item.text().lower():
                match = True
                
            self.table.setRowHidden(row, not match)

    def prepare_export_data(self, data):
        export_df = data.copy()
        if export_df.empty:
            return export_df
        if 'failed_subjects' not in export_df.columns:
            export_df['failed_subjects'] = export_df.apply(self.get_failed_subjects_for_row, axis=1)
        export_df = self.sort_by_roll_no(export_df)
        return export_df

    def build_export_summary(self, data):
        total = len(data)
        pass_count = len(data[data['status'] == 'PASS']) if 'status' in data.columns else 0
        fail_count = len(data[data['status'] == 'FAIL']) if 'status' in data.columns else 0
        avg_percentage = round(data['percentage'].mean(), 2) if 'percentage' in data.columns and total else 0
        top_student = "N/A"
        if total and 'total' in data.columns and 'name' in data.columns:
            top_student = str(data.loc[data['total'].idxmax(), 'name'])
        return [
            ("Total Students", total),
            ("Pass", pass_count),
            ("Fail", fail_count),
            ("Pass %", round((pass_count / total) * 100, 2) if total else 0),
            ("Average %", avg_percentage),
            ("Top Student", top_student),
        ]

    def export_to_excel(self, data, filename="Report.xlsx"):
        if data is None or data.empty:
            QMessageBox.warning(self, "No Data", "No data available to export.")
            return
        
        _default_xlsx = os.path.join(BASE_DIR, filename)
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", _default_xlsx, "Excel Files (*.xlsx)")
        if path:
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            try:
                export_df = self.prepare_export_data(data)
                summary_rows = self.build_export_summary(export_df)
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    export_df.to_excel(writer, index=False, sheet_name="Students")

                from openpyxl import load_workbook
                wb = load_workbook(path)
                ws = wb["Students"]

                header_fill = PatternFill("solid", fgColor="F97316")
                header_font = Font(color="FFFFFF", bold=True)
                thin = Side(style="thin", color="E2E8F0")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                ws.freeze_panes = "A2"
                ws.sheet_view.showGridLines = False
                ws.auto_filter.ref = ws.dimensions
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border

                status_col = None
                failed_subjects_col = None
                for idx, cell in enumerate(ws[1], start=1):
                    if str(cell.value).lower() == "status":
                        status_col = idx
                    if str(cell.value).lower() == "failed_subjects":
                        failed_subjects_col = idx

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        cell.border = border
                    if status_col:
                        status_cell = row[status_col - 1]
                        if status_cell.value == "PASS":
                            status_cell.fill = PatternFill("solid", fgColor="DCFCE7")
                            status_cell.font = Font(color="166534", bold=True)
                        elif status_cell.value == "FAIL":
                            status_cell.fill = PatternFill("solid", fgColor="FEE2E2")
                            status_cell.font = Font(color="991B1B", bold=True)
                    if failed_subjects_col:
                        row[failed_subjects_col - 1].alignment = Alignment(vertical="top", wrap_text=True)

                for column_cells in ws.columns:
                    max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                    width = min(max(max_len + 2, 11), 42)
                    ws.column_dimensions[column_cells[0].column_letter].width = width
                if failed_subjects_col:
                    ws.column_dimensions[ws.cell(1, failed_subjects_col).column_letter].width = 56

                if ws.max_row > 1 and ws.max_column > 1:
                    table = Table(displayName="StudentResultsTable", ref=f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}")
                    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    table.tableStyleInfo = style
                    ws.add_table(table)

                summary = wb.create_sheet("Summary", 0)
                summary.sheet_view.showGridLines = False
                summary["A1"] = "Student Result Report — VNSGU SASCMA"
                summary["A1"].font = Font(size=18, bold=True, color="0C1A3A")
                summary["A2"] = QDate.currentDate().toString("dddd, MMMM dd, yyyy")
                summary["A2"].font = Font(color="64748B")

                for r, (label, value) in enumerate(summary_rows, start=4):
                    summary.cell(r, 1, label)
                    summary.cell(r, 2, value)
                    summary.cell(r, 1).font = Font(bold=True, color="334155")
                    summary.cell(r, 2).font = Font(bold=True, color="0D9488")
                    summary.cell(r, 1).fill = PatternFill("solid", fgColor="F8FAFC")
                    summary.cell(r, 2).fill = PatternFill("solid", fgColor="F0FDFA")
                    summary.cell(r, 1).border = border
                    summary.cell(r, 2).border = border

                chart_row = 13
                summary.cell(chart_row, 1, "Status"); summary.cell(chart_row, 2, "Count")
                summary.cell(chart_row + 1, 1, "PASS"); summary.cell(chart_row + 1, 2, summary_rows[1][1])
                summary.cell(chart_row + 2, 1, "FAIL"); summary.cell(chart_row + 2, 2, summary_rows[2][1])

                pie = PieChart()
                pie.title = "Pass vs Fail"
                pie.add_data(Reference(summary, min_col=2, min_row=chart_row, max_row=chart_row + 2), titles_from_data=True)
                pie.set_categories(Reference(summary, min_col=1, min_row=chart_row + 1, max_row=chart_row + 2))
                pie.height = 7; pie.width = 9
                summary.add_chart(pie, "D4")

                grade_counts = export_df['grade'].value_counts().sort_index() if 'grade' in export_df.columns else pd.Series(dtype=int)
                grade_start = 18
                summary.cell(grade_start, 1, "Grade"); summary.cell(grade_start, 2, "Students")
                for offset, (grade, count) in enumerate(grade_counts.items(), start=1):
                    summary.cell(grade_start + offset, 1, str(grade))
                    summary.cell(grade_start + offset, 2, int(count))

                if len(grade_counts) > 0:
                    bar = BarChart()
                    bar.title = "Grade Distribution"
                    bar.y_axis.title = "Students"
                    bar.x_axis.title = "Grade"
                    bar.add_data(Reference(summary, min_col=2, min_row=grade_start, max_row=grade_start + len(grade_counts)), titles_from_data=True)
                    bar.set_categories(Reference(summary, min_col=1, min_row=grade_start + 1, max_row=grade_start + len(grade_counts)))
                    bar.height = 7; bar.width = 12
                    summary.add_chart(bar, "D20")

                summary.column_dimensions["A"].width = 22
                summary.column_dimensions["B"].width = 18
                wb.save(path)
                QMessageBox.information(self, "Success", f"Data exported to:\n{path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to export: {e}")


    def browse_pdf(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open PDF", "", "PDF Files (*.pdf)")
        if path: self.start_processing(path)

    def drag_enter_event(self, e):
        if e.mimeData().hasUrls(): e.accept()
        else: e.ignore()

    def drop_event(self, e):
        files = [u.toLocalFile() for u in e.mimeData().urls()]
        if files: self.start_processing(files[0])

    def start_processing(self, path):
        if self.worker and self.worker.isRunning(): return
        self.current_pdf_path = path
        self.progress.setValue(0); self.progress.setVisible(True)
        self.browse_btn.setEnabled(False)
        self.status_lbl.setText("⏳ Processing PDF... Please wait.")
        
        self.worker = PDFWorker(path)
        self.worker.progress.connect(self.progress.setValue)
        self.worker.log.connect(self.status_lbl.setText)
        self.worker.finished.connect(self.on_processing_finished)
        self.worker.error.connect(self.on_processing_error)
        self.worker.start()

    def on_processing_finished(self, payload):
        self.progress.setVisible(False)
        self.browse_btn.setEnabled(True)

        try:
            students      = payload["students"]
            subjects_map  = payload["subjects_map"]
            ext_passing   = payload["ext_passing"]
            max_marks     = payload.get("max_marks", 700)

            # Start a new import session in the DB
            session_id = self.db.start_import_session(
                pdf_filename  = payload.get("pdf_filename",  ""),
                course        = payload.get("course",        "Unknown Course"),
                semester      = payload.get("semester",      "Unknown Semester"),
                academic_year = payload.get("academic_year", ""),
                college_name  = payload.get("college_name",  ""),
                max_marks     = max_marks,
                subject_count = len(subjects_map) or 7,
            )

            # Insert all students with per-subject marks
            result = self.db.insert_students(
                students_list  = students,
                session_id     = session_id,
                subjects_map   = subjects_map,
                ext_passing_min= ext_passing,
                max_marks      = max_marks,
            )

            if result is True:
                session = self.db.get_session_info()
                course_txt = (
                    f"{session.get('course','')} | "
                    f"{session.get('semester','')} | "
                    f"{session.get('academic_year','')}"
                ).strip(" |")
                self.status_lbl.setText(
                    f"✅ {len(students)} students loaded"
                    + (f" — {course_txt}" if course_txt else "")
                )
                self.unlock_features()
            else:
                err_msg = str(result) if result else "Unknown database error."
                QMessageBox.critical(self, "Save Error",
                    f"Data processed but failed to save to database:\n{err_msg}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Failed to process result:\n{str(e)}")

    def on_processing_error(self, msg):
        self.progress.setVisible(False); self.browse_btn.setEnabled(True)
        self.status_lbl.setText("❌ Error occurred.")
        QMessageBox.critical(self, "Error", msg)

    def clear_db(self):
        reply = QMessageBox.question(self, "Confirm Reset", "Are you sure you want to reset all data?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.db.delete_all()   # clear all 3 tables, keep connection alive
        self.data_loaded = False
        self.full_df = pd.DataFrame()
        self.current_pdf_path = ""

        for i in range(self.menu_list.count()):
            item = self.menu_list.item(i)
            if item and item.data(Qt.UserRole) != 1:
                item.setFlags(Qt.NoItemFlags)

        self.switch_page(1)
        self.refresh_dashboard()
        self.refresh_failed_page()
        self.refresh_my_functions()
        self.my_ana_canvas.axes.clear()
        self.my_ana_canvas.draw()
        QMessageBox.information(self, "Reset Complete", "Application has been reset. Please upload a new PDF.")

    def refresh_dashboard(self):
        try:
            stats = self.db.get_stats()
            self.full_df = self.sort_by_roll_no(self.db.get_all_data())
            
            if stats:
                self.kpi_total.lbl_value.setText(str(stats["total_students"]))
                self.kpi_pass.lbl_value.setText(f"{stats['pass_perc']}%")
                self.kpi_fail.lbl_value.setText(f"{stats['fail_perc']}%")
                self.kpi_topper.lbl_value.setText(str(stats["topper"]))
            else:
                self.kpi_total.lbl_value.setText("0")
                self.kpi_pass.lbl_value.setText("0%")
                self.kpi_fail.lbl_value.setText("0%")
                self.kpi_topper.lbl_value.setText("N/A")

            self.table.setSortingEnabled(False)
            self.table.clearContents()
            self.table.setRowCount(len(self.full_df))
            for r_idx, (_, r_data) in enumerate(self.full_df.iterrows()):
                items = [str(r_data['roll_no']), str(r_data['name'])]
                for i in range(7):
                    items.append(str(r_data[f'sub_{i+1}']))
                items.extend([str(r_data['total']), str(r_data['sgpa']), str(r_data['status']), str(r_data['atkt_count'])])
                
                for c_idx, txt in enumerate(items):
                    item = QTableWidgetItem(txt)
                    if c_idx == 11:  # Status Column
                        if r_data['status'] == 'FAIL':
                            item.setForeground(QColor("#ffffff"))
                            item.setBackground(QColor("#fee2e2"))
                            item.setForeground(QColor("#dc2626"))
                        else:
                            item.setForeground(QColor("#16a34a"))
                    self.table.setItem(r_idx, c_idx, item)
            
            self.update_charts(self.full_df)
        except Exception as e:
            print(f"Dash Err: {e}")

    def update_charts(self, df):
        if df.empty: return
        try:
            text_color = '#0c1a3a'
            
            self.pie_canvas.fig.patch.set_facecolor('#ffffff')
            self.pie_canvas.axes.clear()
            self.pie_canvas.axes.set_facecolor('#ffffff')
            sc = df['status'].value_counts()
            sc = sc.reindex(['PASS', 'FAIL'], fill_value=0)
            
            self.pie_canvas.axes.pie(
                sc,
                labels=sc.index,
                autopct='%1.1f%%',
                startangle=90,
                colors=['#0d9488', '#f97316'],
                wedgeprops={'linewidth': 3, 'edgecolor': '#ffffff'},
                textprops={'color': text_color, 'fontsize': 11, 'fontweight': 'bold'}
            )
            self.pie_canvas.axes.set_title("Pass / Fail Ratio", color=text_color, fontsize=13, fontweight='bold')
            self.pie_canvas.draw()

            self.bar_canvas.fig.patch.set_facecolor('#ffffff')
            self.bar_canvas.axes.clear()
            self.bar_canvas.axes.set_facecolor('#f8fafc')
            gc = df['grade'].value_counts().sort_index()
            bar_palette = ['#0d9488', '#7c3aed', '#0891b2', '#16a34a', '#ca8a04', '#f97316', '#dc2626']
            chart_colors = [bar_palette[i % len(bar_palette)] for i in range(len(gc))]
            self.bar_canvas.axes.bar(gc.index, gc.values, color=chart_colors, edgecolor='#ffffff', linewidth=1.5, width=0.55)
            self.bar_canvas.axes.set_title("Grade Distribution", color=text_color, fontsize=13, fontweight='bold')
            self.bar_canvas.axes.set_xlabel("Grade", color='#64748b', fontsize=10)
            self.bar_canvas.axes.set_ylabel("No. of Students", color='#64748b', fontsize=10)
            self.bar_canvas.axes.grid(axis='y', color='#e2e8f0', linewidth=0.8)
            self.bar_canvas.axes.tick_params(colors='#64748b', labelsize=10)
            self.bar_canvas.axes.spines['top'].set_visible(False)
            self.bar_canvas.axes.spines['right'].set_visible(False)
            self.bar_canvas.axes.spines['left'].set_color('#e2e8f0')
            self.bar_canvas.axes.spines['bottom'].set_color('#e2e8f0')
            self.bar_canvas.draw()
        except Exception as e:
            print(f"Chart Err: {e}")

    def refresh_my_functions(self, *_args):
        self.update_my_functions_charts()

# ==============================================================================
# 8. ENTRY POINT
# ==============================================================================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setAttribute(Qt.AA_EnableHighDpiScaling)
    app.setAttribute(Qt.AA_UseHighDpiPixmaps)
    if os.path.exists(ICON_PATH): app.setWindowIcon(QIcon(ICON_PATH))
    window = StudentAnalyzerApp()
    sys.exit(app.exec())
